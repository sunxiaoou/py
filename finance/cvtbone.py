#! /usr/bin/python3
import re
import sys
from datetime import datetime
from pprint import pprint

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

pd.set_option('display.max_rows', 200)
pd.set_option('display.max_columns', 100)


def get_bones(xlsx: str) -> (pd.DataFrame, pd.DataFrame):
    wb = load_workbook(xlsx, data_only=True)

    ws = wb['强赎预警']
    dic = {}
    for j in range(1, ws.max_column):
        title = ws.cell(row=1, column=j).value
        if title in ['代码', '强赎天计数']:
            dic[title] = j
            if(len(dic)) == 2:
                break
    dic2 = {}
    for i in range(2, ws.max_row):
        code = ws.cell(row=i, column=dic['代码']).value
        if not isinstance(code, int):
            break
        dic2[str(code)] = ws.cell(row=i, column=dic['强赎天计数']).value

    ws = wb['今天可转债']
    for j in range(1, ws.max_column):
        if not ws.cell(row=2, column=j).value:
            break
    j = 31
    data = []
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row - 2, max_col=j):
        data.append([cell.value for cell in row])
    titles = [i for i in data[0] if i]
    title = titles[0]
    for i in range(1, len(data[0])):
        if data[0][i] is None:
            data[0][i] = title
        else:
            title = data[0][i]
    cols = ['代码', '名称', '价格', '涨幅']
    frame = pd.DataFrame(data[2:], columns=[data[0], data[1]])
    # print(frame)
    n130 = frame[titles.pop()][cols]      # .head(int(sys.argv[2]))
    n130.index = np.arange(1, len(n130) + 1)
    n130['代码'] = n130['代码'].apply(lambda x: str(x))
    n130['强赎天计数'] = n130['代码'].apply(lambda x: dic2[x] if x in dic2 else None)

    df = pd.concat([frame[i][cols] for i in titles]).drop_duplicates()
    df = df[df['价格'] != '#N/A']
    df.index = np.arange(1, len(df) + 1)
    df = df.reindex(axis=1)
    df['代码'] = df['代码'].apply(lambda x: str(x))
    df['无阈值排名'] = df.index
    df['170阈值排名'] = df['无阈值排名'][df['价格'] < 170].rank()
    df['150阈值排名'] = df['170阈值排名'][df['价格'] < 150].rank()
    df['130阈值排名'] = df['150阈值排名'][df['价格'] < 130].rank()
    # print(df.dtypes)
    df['强赎天计数'] = df['代码'].apply(lambda x: dic2[x] if x in dic2 else None)
    df.rename({'170阈值排名': '170排名', '150阈值排名': '150排名', '130阈值排名': '130排名'}, axis=1, inplace=True)
    # print(df)

    return n130, df


def get_my_list(xlsx: str) -> pd.DataFrame:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.worksheets[-1]
    lst = []
    for i in range(1, ws.max_row):
        try:
            code = ws.cell(row=i, column=3).value
            name = ws.cell(row=i, column=4).value
            if code[0] in ['1', '7'] and name[2] == '转':
                lst.append((code, name))
        except TypeError:
            pass
    df = pd.DataFrame(lst, columns=['代码', '名称']).drop_duplicates()
    return df


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError:
        df.to_excel(xlsx, sheet_name=sheet, index=False)
        print(xlsx + ' created')
        return

    ws = wb.copy_worksheet(wb.worksheets[-1])       # copy a old sheet as template to avoid adjust size
    if df.shape[0] < ws.max_row:
        ws.delete_rows(df.shape[0], ws.max_row - 1)
    ws.title = sheet
    wb.active = len(wb.worksheets) - 1

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()


def main():
    if len(sys.argv) < 2:
        print('Usage: {} ref_rank_list.xlsx'.format(sys.argv[0]))
        print('Usage: {} ref_rank_list.xlsx 20'.format(sys.argv[0]))
        sys.exit(1)

    mine = get_my_list('asset.xlsx')

    xlsx = sys.argv[1]
    new130, bones = get_bones(xlsx)
    b2 = bones

    if len(sys.argv) > 2:
        rank130 = int(sys.argv[2])
        new130 = new130.head(rank130)
        print('新版130排名({})'.format(rank130))
        print(new130)
        df = bones.loc[bones['130排名'] == rank130]
        r = df.iloc[0, df.columns.get_loc('无阈值排名')]
        bones = bones[bones['无阈值排名'] <= r]
        print('无阈值排名(截止到<130的第{}名)'.format(rank130))
        print(bones)
        if len(sys.argv) == 4 and sys.argv[3].endswith('.xlsx'):
            to_excel(sys.argv[3], re.search(r'\d{8}', xlsx)[0], bones)

    inner = pd.merge(new130, mine, on=['代码', '名称'])
    print('已持有的新版130转债({})'.format(len(inner)))
    print(inner)

    to_buy = pd.concat([new130, inner]).drop_duplicates(keep=False)
    print('未持有的新版130转债({})'.format(len(to_buy)))
    print(to_buy)

    inner = pd.merge(bones, mine, on=['代码', '名称'])
    print('已持有的激进转债({})'.format(len(inner)))
    print(inner)

    to_buy = pd.concat([bones, inner]).drop_duplicates(keep=False)
    to_buy = to_buy[to_buy['价格'] < 170]
    print('未持有的<170的激进转债({})'.format(len(to_buy)))
    print(to_buy)

    to_sell = pd.concat([inner[['代码', '名称']], mine]).drop_duplicates(keep=False)
    to_sell = pd.merge(to_sell, b2, on=['代码', '名称'], how='left')
    print('持有的非激进转债({})'.format(len(to_sell)))
    if len(to_sell):
        print(to_sell)


if __name__ == "__main__":
    main()
