#! /usr/bin/python3
import os
import re
import sys
from pprint import pprint

import openpyxl
import pandas as pd
import requests
from forex_python.converter import CurrencyRates
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

from securities import *

# pd.set_option('display.max_rows', 1000)
# pd.set_option('display.max_columns', 10)


columns = ['platform', 'currency', 'code', 'name', 'type', 'risk', 'market_value', 'hold_gain']


def verify(row: pd.Series):
    if not pd.isna(row['nav']):
        mv = round(row['nav'] * row['volume'], 2)
        if row['market_value'] != mv:
            print(row['nav'], type(row['nav']))
            print('Warning: {} market_value({}) != {}'.format(row['name'], row['market_value'], mv))
        hg = round((row['nav'] - row['cost']) * row['volume'], 2)
        if abs(row['hold_gain'] - hg) >= 1:
            print('Warning: {} hold_gain({}) != {}'.format(row['name'], row['hold_gain'], hg))


def zhaoshang_bank(datafile: str) -> pd.DataFrame:
    bonds = ['招智睿远平衡二十七期']

    with open(datafile) as fp:
        lines = [re.sub(r'[,，]', '', line).rstrip('\n') for line in fp.readlines()]
    i = 0
    while lines[i] != '尾号8884':
        i += 1
    cash = float(lines[i + 1])
    result = [('招商银行', 'cny', 'cash', '现金', '货币', 0, cash, 0)]
    i += 2
    while lines[i] != '理财产品':
        i += 1
    total_mv = float(lines[i + 1].rstrip('>'))
    asset = round(cash + total_mv, 2)
    i += 2
    try:
        while lines[i]:
            name = lines[i]
            i += 1
            while not re.match(r'.*[\d.]+', lines[i]):
                i += 1
            hold_gain = float(re.sub('[^\d.]+', '', lines[i]))
            market_value = float(lines[i + 1])
            type, risk = ('债券', 1) if name in bonds else ('货币', 0)
            result.append(('招商银行', 'cny', 'product', name, type, risk, market_value, hold_gain))
            i += 4
    except IndexError:
        pass
    df = pd.DataFrame(result, columns=columns)
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    return df


def hangseng_bank(datafile: str) -> pd.DataFrame:
    cash = float(re.sub(r'.+_', '', datafile[: -4]))
    result = [('恒生银行', 'cny', 'cash', '现金', '货币', 0, cash, 0)]

    with open(datafile) as fp:
        lines = [re.sub(r'[,＋]', '', re.sub('－', '-', line)).rstrip('\n') for line in fp.readlines()]
    i = 0
    while lines[i] != '总市值（元）':
        i += 1
    total_mv = float(lines[i + 1])
    asset = round(cash + total_mv, 2)
    total_hg = float(re.sub(r'[^-\d.]+', '', lines[i + 2]))
    i += 3
    # print(cash, total_mv, total_hg)
    while lines[i] != '交易记录':
        i += 1
    i += 1
    while re.match(r'\d{6}.*', lines[i]):
        if len(lines[i]) == 6:
            code = lines[i]
            i += 1
        else:
            code = lines[i][: 6]
        name, type, risk = off_market[code]
        i += 1
        hold_gain = float(lines[i])
        i += 6
        if re.match(r'.*[\d.]+', lines[i]):
            market_value = float(re.sub(r'[^\d.]+', '', lines[i]))
        else:
            i += 1
            market_value = float(lines[i])
        result.append(('恒生银行', 'cny', code, name, type, risk, market_value, hold_gain))
        i += 1
    df = pd.DataFrame(result, columns=columns)
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    return df


def yinhe(datafile: str) -> pd.DataFrame:
    with open(datafile) as fp:
        lines = [re.sub('－', '-', line).rstrip('\n') for line in fp.readlines()]

    i = 0
    while lines[i] != '场内资产（人民币）':
        i += 1
    asset = float(lines[i + 1])
    total_mv = float(lines[i + 3])
    total_hg = float(lines[i + 5])
    cash = float(lines[i + 7])
    assert round(total_mv + cash, 2) == asset,\
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('银河', 'cny', 'cash', '现金', '货币', 0, cash, 0)]
    i += 8
    while not lines[i].startswith('参考盈亏'):
        i += 1
    i += 1
    while True:
        while i < len(lines) and not re.match(r".*\d{6}$", lines[i]):
            i += 1
        if i == len(lines):
            break
        if len(lines[i]) == 6:
            name, code = lines[i - 1], lines[i]
        else:
            name, code = lines[i][: -6].rstrip(), lines[i][-6:]
        i += 1
        if code[0] in ['1', '7']:
            type, risk = '转债', 2
        elif code[0] == '2':
            type, risk = '货币', 0
        else:
            name, type, risk = on_market[code]

        hold_gain = float(lines[i])
        i += 1
        if ' ' in lines[i]:     # volume, cost are in same line
            l = lines[i].split()
            volume, cost = int(l[0]), float(l[1])
            i += 1
        else:
            volume, cost = int(lines[i]), float(lines[i + 1])
            i += 2
        market_value = float(lines[i])
        i += 2
        if ' ' in lines[i]:
            l = lines[i].split()
            v2, nav = int(l[0]), float(l[1])
            i += 1
        else:
            v2, nav = int(lines[i]), float(lines[i + 1])
            i += 2
        assert v2 == volume, print("v2{} != volume{}".format(v2, volume))
        assert market_value == round(nav * volume, 2),\
            print("mv({}) != nav({}) * volume({})".format(market_value, nav, volume))
        result.append(('银河', 'cny', code, name, type, risk, market_value, hold_gain, volume, nav, cost))
    df = pd.DataFrame(result, columns=columns + ['volume', 'nav', 'cost'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.apply(verify, axis=1)
    df.drop(columns=['volume', 'cost'], inplace=True)
    return df


def huabao(datafile: str) -> pd.DataFrame:
    with open(datafile) as f:
        lines = []
        for l in f.readlines():
            # lines += re.sub('－', '-', l.rstrip('\n')).split()
            lines += l.rstrip('\n').split()
    i = 0
    while not re.match(r'\d+\.\d\d', lines[i]):
        i += 1
    asset = float(lines[i])
    total_mv = float(lines[i + 1])
    i += 2
    while lines[i] != '可取资金':
        i += 1
    cash = float(lines[i + 1])
    assert round(total_mv + cash, 2) == asset, \
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('华宝', 'cny', 'cash', '现金', '货币', 0, cash, 0)]
    while not lines[i].startswith('今日盈亏'):
        i += 1
    s = lines[i + 2]
    total_hg = float(re.sub('－', '-', s) if s[0] == '－' else s[1:])
    i += 3
    while lines[i] != '仓位':
        i += 1
    i += 1
    try:
        while True:
            while not re.match(r'\d{6}\.(SZ|SH)', lines[i]):
                i += 1
            name = lines[i - 4]
            cost = float(lines[i - 3])
            volume = int(lines[i - 2])
            s = lines[i - 1]
            hold_gain = float(re.sub('－', '-', s) if s[0] == '－' else s[1:])
            code = lines[i][: 6]
            market_value = float(lines[i + 2])
            nav = float(lines[i + 3])
            v2 = int(lines[i + 4])
            assert v2 == volume, print("v2{} != volume{}".format(v2, volume))
            assert market_value == round(nav * volume, 2) or market_value == round(nav * volume * 10, 2), \
                print("mv({}) != nav({}) * volume({})".format(market_value, nav, volume))
            if market_value == round(nav * volume * 10, 2):
                volume *= 10
            if code[0] == '1':
                typ, risk = '转债', 2
            else:
                name, typ, risk = on_market[code]
            i += 6
            result.append(('华宝', 'cny', code, name, typ, risk, market_value, hold_gain, volume, nav, cost))
    except IndexError:
        pass
    df = pd.DataFrame(result, columns=columns + ['volume', 'nav', 'cost'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.apply(verify, axis=1)
    df.drop(columns=['volume', 'cost'], inplace=True)
    return df


def huasheng(datafile: str) -> pd.DataFrame:
    with open(datafile) as fp:
        lines = [re.sub(r'[,＋]', '', re.sub('－', '-', line)).rstrip('\n') for line in fp.readlines()]

    i = 0
    while not lines[i].startswith('资产净值'):
        i += 1
    currency = 'hkd' if '港币' in lines[i] else 'usd'
    asset = float(lines[i + 3])
    total_mv = float(lines[i + 8])
    total_hg = float(lines[i + 10])
    cash = float(lines[i + 14])
    assert round(total_mv + cash, 2) == asset,\
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('华盛', currency, 'cash', '现金', '货币', 0, cash, 0)]
    i += 15
    while not lines[i].startswith('持仓盈亏'):
        i += 1
    i += 1
    while len(lines) - i >= 8:
        name = lines[i]
        volume = int(lines[i + 1])
        hold_gain = float(lines[i + 2])
        nav = float(lines[i + 3])
        code = lines[i + 4].rstrip('融')
        name, type, risk = on_market[code]
        market_value = float(lines[i + 5])
        i += 6
        while not re.match(r'.*[\d.]+$', lines[i]):
            i += 1
        cost = float(lines[i])
        i += 1
        result.append(('华盛', currency, code, name, type, risk, market_value, hold_gain, volume, nav, cost))
    df = pd.DataFrame(result, columns=columns + ['volume', 'nav', 'cost'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.apply(verify, axis=1)
    df.drop(columns=['volume', 'cost'], inplace=True)
    return df


def futu(datafile: str) -> pd.DataFrame:
    with open(datafile) as fp:
        lines = [re.sub(r'[,，＋]', '', re.sub('－', '-', line)).rstrip('\n') for line in fp.readlines()]

    i = 0
    while not lines[i].startswith('资产净值'):
        i += 1
    currency = 'hkd' if '港币' in lines[i] else 'usd'
    asset = float(lines[i + 1])
    i += 2
    while not lines[i].startswith('持仓盈亏'):
        i += 1
    total_hg = float(lines[i + 3])
    i += 4
    while not lines[i].startswith('现金可提'):
        i += 1
    total_mv = float(lines[i + 1])
    cash = float(lines[i + 2])
    i += 3
    assert round(total_mv + cash, 2) == asset, \
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('富途', currency, 'cash', '现金', '货币', 0, cash, 0)]
    i += 2
    while not lines[i].startswith('今日盈亏'):
        i += 1
    i += 1
    while len(lines) - i >= 7:
        code = lines[i + 4]
        volume = lines[i + 5]
        # name = lines[i]
        name, type, risk = on_market[code]
        market_value = float(lines[i + 1])
        hold_gain = float(lines[i + 2])
        nav = round(market_value / int(volume), 2)
        result.append(('富途', currency, code, name, type, risk, market_value, hold_gain, nav))
        i += 7
    df = pd.DataFrame(result, columns=columns + ['nav'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    return df


def danjuan(datafile: str) -> pd.DataFrame:

    def get_plan(plan: str) -> list:
        resp = requests.get(url + 'plan/' + plan, headers=headers)
        assert resp.status_code == 200, print('status_code({}) != 200'.format(resp.status_code))
        result = []
        for i in resp.json()['data']['items']:
            plan_code = i['plan_code']
            code = i['fd_code']
            market_value = float(i['market_value'])
            hold_gain = float(i['hold_gain'])
            nav = float(i['nav'])
            if market_value:
                name, type, risk = off_market[code]
                result.append(('蛋卷' + plan_code, 'cny', code, name, type, risk, market_value, hold_gain, nav))
        return result

    if os.path.isfile(datafile):
        return pd.read_csv(datafile, index_col=0)

    with open('auth/dj_cookie.txt', 'r') as f:
        cookie = f.read()[:-1]      # delete last '\n'
    url = 'https://danjuanapp.com/djapi/holding/'
    headers = {
        'Cookie': cookie,
        'Host': url.split('/')[2],
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_1) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/88.0.4324.192 Safari/537.36'}

    response = requests.get(url + 'summary', headers=headers)
    assert response.status_code == 200, print('status_code({}) != 200'.format(response.status_code))
    asset = float(response.json()['data']['total_assets'])
    total_hg = float(response.json()['data']['hold_gain'])
    result = []
    for i in response.json()['data']['items']:
        code = i['fd_code']
        if code in ['CSI666', 'TIA06020']:       # 螺丝钉指数基金组合, 螺丝钉金钉宝主动优选
            result.extend(get_plan(code))
        else:
            if code in ['CSI1014', 'TIA06019']:  # 我要稳稳的幸福, 螺丝钉银钉宝365天
                name, type, risk = i['fd_name'], '债券', 1
            else:
                name, type, risk = off_market[code]
            market_value = float(i['market_value'])
            hold_gain = float(i['hold_gain'])
            nav = float(i['nav'])
            result.append(('蛋卷', 'cny', code, name, type, risk, market_value, hold_gain, nav))
    df = pd.DataFrame(result, columns=columns + ['nav'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert abs(sum_mv - asset) <= 1, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert abs(sum_hg - total_hg) <= 1, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.to_csv(datafile)
    return df


def tonghs(datafile: str) -> pd.DataFrame:
    if os.path.isfile(datafile):
        df = pd.read_csv(datafile, index_col=0)
        df['code'] = df['code'].apply(lambda x: '{0:06d}'.format(x))
        return df

    with open('auth/ths_cookie.txt', 'r') as f:
        cookie = f.read()[:-1]      # delete last '\n'
    url = 'https://trade.5ifund.com/pc_query/trade_queryIncomeWjZeroList.action?_=1615356614458'
    url2 = 'https://trade.5ifund.com/pc_query/trade_currentShareList.action?_=1615344156640'
    headers = {
        'Cookie': cookie,
        'Host': url.split('/')[2],
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_1) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/88.0.4324.192 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest'}

    result = []
    response = requests.get(url, headers=headers)
    assert response.status_code == 200, print('status_code({}) != 200'.format(response.status_code))
    items = response.json()['singleData']['IncomeShareListResult']
    for i in items:
        code = i['fundCode']
        name, type, risk = off_market[code]
        market_value = float(i['totalVol'])
        hold_gain = float(i['sumIncome'])
        nav = 1.0
        result.append(('同花顺', 'cny', code, name, type, risk, market_value, hold_gain, nav))
    response = requests.get(url2, headers=headers)
    assert response.status_code == 200, print('status_code({}) != 200'.format(response.status_code))
    items = response.json()['singleData']['currentShareList']
    for i in items:
        code = i['fundCode']
        name, type, risk = off_market[code]
        market_value = float(i['currentValueText'])
        hold_gain = float(i['totalprofitlossText'])
        nav = float(i['navText'])
        result.append(('同花顺', 'cny', code, name, type, risk, market_value, hold_gain, nav))
    df = pd.DataFrame(result, columns=columns + ['nav'])
    df.to_csv(datafile)
    return df


def run(file: str) -> pd.DataFrame:
    platforms = {'zsb': zhaoshang_bank, 'hsb': hangseng_bank, 'yh': yinhe, 'hb': huabao, 'hs': huasheng, 'ft': futu,
                 'dj': danjuan, 'ths': tonghs}
    func = platforms[re.search(r'.*/(.*)_.*', file).group(1)]
    return func(file)


def to_cny(row: pd.Series, key: str, rates: tuple) -> float:
    h2c, u2c = rates
    value = row[key] * h2c if row['currency'] == 'hkd' else row[key] * u2c if row['currency'] == 'usd' else row[key]
    return round(value, 2)


def gain_rate(row: pd.Series) -> float:
    try:
        rate = row['hold_gain'] / (row['market_value'] - row['hold_gain'])
    except ZeroDivisionError:
        rate = 0
    # return '{:.2%}'.format(rate)
    return round(rate, 4)


def fill(df: pd.DataFrame) -> pd.DataFrame:
    cr = CurrencyRates()
    h2c = round(cr.get_rate('HKD', 'CNY'), 2)
    u2c = round(cr.get_rate('USD', 'CNY'), 2)
    # print(h2c, u2c)
    df['name'] = df['name'].apply(lambda s: s if len(s) <= 10 else s[: 8] + '..')   # truncate name
    df['market_value'] = df.apply(lambda row: to_cny(row, 'market_value', (h2c, u2c)), axis=1)
    df['hold_gain'] = df.apply(lambda row: to_cny(row, 'hold_gain', (h2c, u2c)), axis=1)
    df['gain_rate'] = df.apply(gain_rate, axis=1)
    col2 = ['platform', 'currency', 'code', 'name', 'type', 'risk', 'nav', 'market_value', 'hold_gain', 'gain_rate']
    return df.reindex(columns=col2)


def run_all(files: list) -> pd.DataFrame:
    platforms = ['zsb_', 'hsb_', 'yh_', 'hb_', 'hs_', 'ft_', 'dj_', 'ths_']   # to sort files
    fs = sorted(files)
    frames = []
    for p in platforms:
        for f in fs:
            if os.path.basename(f).startswith(p):
                # print(f, p)
                frames.append(run(f))
    df = pd.concat(frames)
    df = fill(df)
    return df


def to_execl(xlsx: str, sheet: str, df: pd.DataFrame):
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()

    ws = wb[sheet]
    last_row = ws.max_row
    last_col = ws.max_column

    for i in range(2, last_row + 1):
        for j in range(6, last_col):
            ws.cell(row=i, column=j).number_format = '#,##,0.00'
        ws.cell(row=i, column=last_col).number_format = '0.00%'

    summaries = [
        {'location': (last_row + 2, 1),
         'letter': 'A',
         'labels': ['招商银行', '恒生银行', '银河', '华宝', '华盛*', '富途*', '蛋卷*', '同花顺'],
         'category': 'platform',
         'anchor': 'K1'},
        {'location': (last_row + 2, 4),
         'letter': 'B',
         'labels': ['cny', 'hkd', 'usd'],
         'category': 'currency',
         'anchor': 'K16'},
        {'location': (last_row + 6, 4),
         'letter': 'F',
         'labels': [0, 1, 2, 3],
         'category': 'risk',
         'anchor': 'K31'},
        {'location': (last_row + 2, 7),
         'letter': 'E',
         'labels': ['货币', '债券', '转债', '指数', '主动', '股票', '医药', '中概股'],
         'category': 'type',
         'anchor': 'K46'}
    ]

    for summary in summaries:
        row, col = summary['location']
        le = [get_column_letter(j) for j in range(col, col + 3)]
        for i in range(len(summary['labels'])):
            ws.cell(row=row+i, column=col).value = summary['labels'][i]
            c = ws.cell(row=row+i, column=col+1)
            c.number_format = "#,##,0.00"
            c.value = '=SUMIF(${0}$2:${0}${1},{2}{3},$H$2:$H${1})'.format(summary['letter'],
                                                                          last_row, le[0], row + i)
            c = ws.cell(row=row+i, column=col+2)
            c.number_format = "#,##,0.00"
            c.value = '=SUMIF(${0}$2:${0}${1},{2}{3},$I$2:$I${1})'.format(summary['letter'],
                                                                          last_row, le[0], row + i)
        ws.cell(row=row+i+1, column=col).value = 'sum'
        c = ws.cell(row=row+i+1, column=col+1)
        c.number_format = "#,##,0.00"
        c.value = '=SUM({0}{1}:{0}{2})'.format(le[1], row, row + i)
        c = ws.cell(row=row+i+1, column=col+2)
        c.number_format = "#,##,0.00"
        c.value = '=SUM({0}{1}:{0}{2})'.format(le[2], row, row + i)

        if summary['letter'] == 'E':
            for j in range(len(summary['labels']) + 1):
                c = ws.cell(row=row+j, column=col+3)
                c.number_format = '0.00%'
                c.value = '={1}{0}/({2}{0}-{1}{0})'.format(row + j, le[2], le[1])

        pie = PieChart()
        labels = Reference(ws, min_col=col, min_row=row, max_row=row+i)
        data = Reference(ws, min_col=col+1, min_row=row-1, max_row=row+i)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = summary['category']
        ws.add_chart(pie, summary['anchor'])
    wb.save(xlsx)


def main():
    if len(sys.argv) < 2:
        print('Usage: {} txt'.format(sys.argv[0]))
        print('       {} dir [xlsx]'.format(sys.argv[0]))
        sys.exit(1)

    path = sys.argv[1]
    if os.path.isfile(path) or path.endswith('.csv'):
        print(run(path))
        sys.exit(0)

    assert(os.path.isdir(path))
    df = run_all([os.path.join(path, file) for file in os.listdir(path)])

    if len(sys.argv) == 2:
        print(df)
        sys.exit(0)

    to_execl(sys.argv[2], path, df)


if __name__ == "__main__":
    main()
