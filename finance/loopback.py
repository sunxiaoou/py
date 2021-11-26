#! /usr/bin/python3
import re
import sys
from datetime import date
from datetime import datetime

import matplotlib.pyplot as plt
import numpy
import openpyxl
import pandas as pd
import akshare as ak
from openpyxl import load_workbook

# pd.set_option('display.max_rows', 1000)
# pd.set_option('display.max_columns', 6)

names = {
    "sz159928": "消费ETF",
    "sz161039": "1000增强",
    "sz164906": "中概互联网LOF",
    "sh501009": "生物科技LOF",
    "sh501050": "50AH",
    "sh501090": "消费龙头LOF",
    "sh502000": "500增强LOF",
    "sh510310": "沪深300",
    # "sh510580": "中证500ETF易方达",
    "sh510710": "上证50",
    "sh512000": "券商ETF",
    # "sh512170": "医疗ETF",
    "sh512260": "中证500低波ETF",
    "sh512800": "银行ETF",
    "sh513050": "中国互联网ETF",
    "sh515180": "红利ETF易方达",

    "000248": "汇添富中证消费ETF联接",
    "001556": "天弘中证500A",
    "001594": "天弘中证银行A",
    "003318": "景顺500低波",
    "004069": "南方中证全指证券",
    "006327": "易方达中证海外50ETF联接",
    "090010": "大成中证红利",
    "110003": "易方达上证50指数A",
    "162412": "华宝医疗ETF联接A",
    "163407": "兴全沪深300A",
    "164906": "交银中证海外中国互联网",
    "501009": "汇添富中证生物科技",
    "501050": "华夏上证50AH",
    "501090": "华宝中证消费龙头",
    "519671": "银河300价值",
    "540012": "汇丰晋信恒生A股龙头",

    "001643": "汇丰晋信智造先锋",
    "001717": "工银前沿医疗",
    "004868": "交银股息优化",
    "000595": "嘉实泰和",
    "001766": "上投摩根医疗健康",
    "001810": "中欧潜力价值",
    "001974": "景顺长城量化新动力",
    "003095": "中欧医疗健康A",
    "005259": "建信龙头企业",
    "005267": "嘉实价值精选",
    "005354": "富国沪港深行业精选A",
    "006228": "中欧医疗创新A",
    "110011": "易方达优质精选",
    "161005": "富国天惠成长",
    "163402": "兴全趋势投资",
    "163406": "兴全合润",
    "163415": "兴全商业模式优选",
    "166002": "中欧新蓝筹A",
    "169101": "东方红睿丰",
    "377240": "上投摩根新兴动力",
    "519035": "富国天博创新",
    "519688": "交银精选",
    "540003": "汇丰晋信动态策略A"
}


def get_weekly(code: str, begin: date) -> pd.DataFrame:
    base = ak.stock_zh_index_daily(symbol='sh510310')[['date']]     # use '沪深300ETF' as base
    base['date'] = pd.to_datetime(base['date'])
    base = base[base['date'] > begin]

    df = pd.DataFrame()
    if re.match(r'(sh|sz)\d{6}', code) is not None:
        # df = ak.stock_zh_index_daily_tx(symbol=code)[['date', 'close']]
        df = ak.stock_zh_index_daily(symbol=code)[['date', 'close']]
    elif re.match(r'\d{6}', code) is not None:
        df = ak.fund_em_open_fund_info(fund=code, indicator='累计净值走势')   # [['净值日期', '单位净值']]
        df = df.rename({'净值日期': 'date', '累计净值': 'close'}, axis=1)
        if df.dtypes['close'] != numpy.dtype('float64'):
            try:
                df['close'] = df['close'].apply(lambda x: float(x))
            except TypeError:
                pass
    else:
        assert True
    df['date'] = pd.to_datetime(df['date'])
    df = pd.merge(base, df, on='date', how='outer')
    df.fillna(method='ffill', inplace=True)     # fill NaN with previous value
    df.fillna({'close': 1.0}, inplace=True)     # no previous value, fill with 1.0
    weekday = 1                                 # choose Tuesday, Mon: 0, Tue: 1, ... Sun: 6
    df = df[(df['date'] > begin) & (df['date'].dt.dayofweek == weekday)]
    # print(df)
    return df


def get_index_name(code: str) -> str:
    df = ak.stock_zh_index_spot()
    i = df.loc[df['代码'] == code].index[0]
    return df.at[i, '名称']


def xirr(df: pd.DataFrame, date_column: str, amount_column: str) -> float:
    residual = 1
    step = 0.05
    guess = 0.05
    epsilon = 0.0001
    limit = 10000

    df = df.sort_values(by=date_column).reset_index(drop=True)
    df['years'] = df.apply(lambda x: (x[date_column] - df[date_column][0]).days / 365, axis=1)
    while abs(residual) > epsilon and limit > 0:
        limit -= 1
        residual = df.apply(lambda x: x[amount_column] / pow(guess, x['years']), axis=1).sum()
        if abs(residual) > epsilon:
            if residual > 0:
                guess += step
            else:
                guess -= step
                step /= 2.0
    return guess - 1


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()


def loop_back(code: str, begin: date) -> tuple:
    df = get_weekly(code, begin)
    df['每期定投金额'] = 1000
    df['累计定投金额'] = df['每期定投金额'].cumsum()
    fee_rate = 0.001
    df['每期持仓数量'] = df['每期定投金额'] / df['close'] * (1 - fee_rate)
    df['累计持仓数量'] = df['每期持仓数量'].cumsum()
    df['累计持仓净值'] = df['close'] * df['累计持仓数量']
    df['现金流'] = - df['每期定投金额']
    m = df.shape[0]     # number of row
    cumulative_amount = df.iloc[m - 1, df.columns.get_loc('累计定投金额')]
    cumulative_net = df.iloc[m - 1, df.columns.get_loc('累计持仓净值')]
    hold_gain = cumulative_net - cumulative_amount
    df.iloc[m - 1, df.columns.get_loc('现金流')] += cumulative_net
    return_rate = xirr(df, 'date', '现金流')
    try:
        name = '{}({})'.format(names[code], code)
    except KeyError:
        name = code
    df = df.rename({'累计持仓净值': name}, axis=1).set_index('date')
    df.index.name = None
    # print(df)
    return (name, cumulative_amount, cumulative_net, hold_gain, return_rate), df[['累计定投金额', name]]


def main():
    if len(sys.argv) < 3:
        print('Usage: {} code YYYYmmdd'.format(sys.argv[0]))
        sys.exit(1)

    begin = datetime.strptime(sys.argv[2], '%Y%m%d')    # .date()
    # 宽基指数
    codes = ["sz161039", "sh502000", "sh510310", "sh510710", "sh501050", "003318", "090010", "519671"]
    # 行业指数
    # codes = ["000248", "162412", "501009", "sz164906", "sh512000", "sh512800"]
    # 主动基金
    codes = ["001643", "001717", "001810", "005267", "161005", "163402"]

    results = [loop_back(i, begin) for i in codes]
    columns = ['name', '累计定投金额(万)', '累计持仓净值(万)', '累计收益(万)', '内部收益率(%)']
    df = pd.DataFrame([r[0] for r in results], columns=columns)
    df[columns[1]] = df[columns[1]].apply(lambda x: round(x / 10000))
    df[columns[2]] = df[columns[2]].apply(lambda x: round(x / 10000))
    df[columns[3]] = df[columns[3]].apply(lambda x: round(x / 10000))
    df[columns[4]] = df[columns[4]].apply(lambda x: round(x * 100))
    df.set_index('name', inplace=True)
    df.index.name = None
    print(df)
    df2 = pd.concat([r[1] for r in results], axis=1)
    df2 = df2.loc[:, ~df2.columns.duplicated()]        # remove duplicated '累计定投金额'
    print(df2)

    _, axes = plt.subplots(nrows=2, ncols=1)
    typ = '宽基指数'
    # typ = '行业指数'
    typ = '主动基金'
    title = typ + '定投回测：累计持仓曲线及收益率'
    df2.plot(ax=axes[0], figsize=(12, 8), grid=True, rot=0, title=title)
    ax = df.plot.bar(ax=axes[1], figsize=(12, 8), rot=20)
    for i in ax.containers:
        ax.bar_label(i)
    plt.figtext(.9, .1, '- 同光和尘')
    plt.show()


if __name__ == "__main__":
    main()
