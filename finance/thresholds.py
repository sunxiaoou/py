#! /usr/bin/python3
import re
import sys
from pprint import pprint

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

columns = ['代码', '名称', '参考指标', '最低', '低估', '高估', '最高']
thresholds = [
    ('000015', '上证红利', '盈利收益率', 17.5, 10, 6.4, 2.27),
    ('950090', '50AH优选', '盈利收益率', 16.6, 10, 6.4, 2.22),
    ('000925', '基本面50', '盈利收益率', 16.6, 10, 6.4, 2.22),
    ('399550', '央视50', '盈利收益率', 16.6, 10, 6.4, 2.5),
    ('000922', '中证红利', '盈利收益率', 16.6, 10, 6.4, 1.6),
    ('000919', '300价值', '盈利收益率', 16.6, 10, 6.4, 2.22),
    ('000016', '上证50', '盈利收益率', 14.5, 10, 6.4, 2.22),
    ('000010', '上证180', '盈利收益率', 13.9, 10, 6.4, 2.17),
    ('HSCEI', 'H股指数', '盈利收益率', 17.5, 10, 6.4, 3.45),
    ('HSI', '恒生指数', '盈利收益率', 14.5, 10, 6.4, 4.76),

    ('399986', '银行行业', '市净率', 0.75, 0.9, 1.15, 1.4),
    ('399393', '地产行业', '市净率', 1.2, 1.6, 2.2, 4),
    ('399975', '证券行业', '市净率', 1.05, 1.6, 2.2, 4.8),
    ('399967', '军工行业', '市净率', 2.1, 2.6, 4, 9.2),
    ('000827', '环保行业', '市净率', 1.82, 2.3, 3, 5.9),
    ('399995', '基建行业', '市净率', 0.92, 1.1, 1.8, 3.6),
    ('931009', '建筑材料', '市净率', 1.4, 1.8, 2.1, 3.1),

    # ('000985', '中证全指', '市盈率', 11, 14, 21, 54),
    ('CSPSADRP', '红利机会', '市盈率', 8, 13, 20, 30),
    ('930782', '500低波动', '市盈率', 17, 24, 30, 60),
    ('000905', '500增强', '市盈率', 17, 25, 40, 93),
    ('000300', '沪深300', '市盈率', 8, 11, 17, 49),
    ('399812', '中证养老', '市盈率', 17, 21, 27, 36),
    ('399006', '创业板', '市盈率', 27, 25, 45, 138),
    ('399330', '深证100', '市盈率', 12, 18, 24, 64),
    ('000978', '医药100', '市盈率', 23, 28, 36, 63),
    ('399989', '中证医疗', '市盈率', 32, 55, 75, 140),
    ('930743', '生物科技', '市盈率', 33, 57, 81, 135),
    ('000932', '中证消费', '市盈率', 17, 30, 40, 53),
    ('399997', '中证白酒', '市盈率', 15, 30, 40, 71),
    ('930653', '食品饮料', '市盈率', 18, 30, 40, 65),
    ('000989', '可选消费', '市盈率', 15, 18, 26, 45),
    # ('686000', '可选消费', '市净率', 1.7, 2.2, 4, 5.3),
    ('H30094', '消费红利', '市盈率', 11, 25, 33, 45),
    ('931068', '消费龙头', '市盈率', 16, 24, 32, 45),
    ('931357', '沪港深消费50', '市盈率', 25, 33, 42, 56),
    ('399001', '深证成指', '市盈率', 11, 15, 30, 62),
    ('399701', '基本面60', '市盈率', 12, 17, 20, 60),
    ('399702', '基本面120', '市盈率', 13, 18, 22, 60),
    ('399324', '深红利', '市盈率', 11, 15, 22, 44),
    ('NDX', '纳斯达克100', '市盈率', 15, 20, 30, 85),
    ('SPX', '标普500', '市盈率', 5.8, 15, 25, 44),
    ('S5INFT', '标普科技', '市盈率', 15, 21, 30, 90),
    ('IXY', '美股消费', '市盈率', 15, 21, 30, 45),
    ('SPG120035', '全球医疗', '市盈率', 15, 21, 30, 45),
    ('SPHCMSHP', '香港中小', '市盈率', 8.4, 12, 17, 20),
    ('H30533', '中概互联', '市销率', 3.74, 5.6, 8, 12.8),
    ('HSCAIT', 'A股龙头', '市盈率', 8, 11, 13, 15),
    ('931142', '竞争力指数', '市盈率', 10, 13, 18, 23),
    ('707717', 'MSCI质量', '市盈率', 17, 26, 38, 55),
    ('000688', '科创50', '市盈率', 55, 50, 80, 100),
    ('930697', '家用电器', '市盈率', 12, 17, 20, 28)]


def exchange():
    for i in thresholds:
        name, code, reference, low, high, highest, lowest = i
        print('(\'{}\', \'{}\', \'{}\', {}, {}, {}, {}),'.format(code, name, reference, lowest, low, high, highest))


def is_number(s: str) -> bool:
    if re.match("^\d+\.?\d+?$", s) is None:
        return s.isdigit()
    return True


def parse(l: list) -> list:
    date = l[3]
    valuations = []
    i = 0
    while l[i] != date:
        i += 1
    while True:
        while l[i] != "场外代码":
            i += 1
        i += 1
        while True:
            k = l[i]
            if k == '美股优秀行业进入绿色是进入到估值中枢以下':
                return valuations
            i += 1
            while not is_number(l[i]):
                k += l[i]
                i += 1
            if k.startswith('中概互联'):
                k = '中概互联'
            v = l[i]
            i += 1
            if v == date:
                break
            valuations.append((k, float(v)))
            while is_number(l[i]):
                i += 1


def to_lowest(row: pd.Series) -> float:
    val, lowest, low = row['当日'], row['最低'], row['低估']
    if row['参考指标'] == '盈利收益率':
        if val < low:
            return None
    elif val > low:
        return None
    return round((lowest - val) / (lowest - low), 4)


def to_lowest2(row: pd.Series) -> float:
    val, lowest, low = row['当日'], row['最低'], row['低估']
    if row['参考指标'] == '盈利收益率':
        if val < low:
            return None
        return round((val - lowest) / val, 4)
    if val > low:
        return None
    return round((lowest - val) / val, 4)


def to_low(row: pd.Series) -> float:
    val, low, high = row['当日'], row['低估'], row['高估']
    if row['参考指标'] == '盈利收益率':
        if val >= low or val <= high:
            return None
    elif val <= low or val >= high:
        return None
    return round((low - val) / (low - high), 4)


def to_low2(row: pd.Series) -> float:
    val, low, high = row['当日'], row['低估'], row['高估']
    if row['参考指标'] == '盈利收益率':
        if val <= high:
            return None
        return round((val - low) / val, 4)
    if val >= high:
        return None
    return round((low - val) / val, 4)


def to_high(row: pd.Series) -> float:
    val, high, highest = row['当日'], row['高估'], row['最高']
    if row['参考指标'] == '盈利收益率':
        if val > high:
            return None
    elif val < high:
        return None
    return round((high - val) / (high - highest), 4)


def to_high2(row: pd.Series) -> float:
    val, low, high = row['当日'], row['低估'], row['高估']
    if row['参考指标'] == '盈利收益率':
        if val >= low:
            return None
        return round((val - high) / val, 4)
    if val <= low:
        return None
    return round((high - val) / val, 4)


def to_highest2(row: pd.Series) -> float:
    val, high, highest = row['当日'], row['高估'], row['最高']
    if row['参考指标'] == '盈利收益率':
        if val >= high:
            return None
        return round((val - highest) / val, 4)
    if val <= high:
        return None
    return round((highest - val) / val, 4)


def calculate_thresholds(vals: list) -> pd.DataFrame:
    df = pd.DataFrame(thresholds, columns=columns)
    # print(df)
    df2 = pd.DataFrame(vals, columns=['名称', '当日'])
    # print(df2)
    df = pd.merge(df, df2, on='名称')
    df['距最低'] = df.apply(to_lowest, axis=1)     # to_lowest(), to_low(), to_high() are just for sort
    df['距低估'] = df.apply(to_low, axis=1)
    df['距高估'] = df.apply(to_high, axis=1)
    df = df.sort_values(by=['距最低', '距低估', '距高估'])
    df['距最低'] = df.apply(to_lowest2, axis=1)
    df['距低估'] = df.apply(to_low2, axis=1)
    df['距高估'] = df.apply(to_high2, axis=1)
    df['距最高'] = df.apply(to_highest2, axis=1)
    return df.reset_index(drop=True)


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    wb = load_workbook(xlsx)
    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()

    wb = load_workbook(xlsx)
    ws = wb[sheet]
    last_row = ws.max_row
    last_col = ws.max_column
    for i in range(2, last_row + 1):
        for j in range(4, 9):
            ws.cell(row=i, column=j).number_format = '#,##,0.00'
        for j in range(9, last_col + 1):
            ws.cell(row=i, column=j).number_format = '0.00%'

    colors = ['99CC00', 'FFCC00', 'FF6600']
    i = 1
    for k in range(9, 12):
        fill = PatternFill(patternType='solid', fgColor=colors[k - 9])
        while ws.cell(row=i, column=k).value is not None:
            for j in range(1, last_col + 1):
                ws.cell(row=i, column=j).fill = fill
            i += 1

    wb.save(xlsx)


def main():
    if len(sys.argv) < 2:
        print('Usage: {} txt'.format(sys.argv[0]))
        print('       {} txt xlsx'.format(sys.argv[0]))
        sys.exit(1)

    with open(sys.argv[1]) as fp:
        lines = [line.rstrip('%\n') for line in fp.readlines()]

    date = lines[3]
    df = calculate_thresholds(parse(lines))
    df.rename(columns={'当日': date}, inplace=True)
    print(df)

    if len(sys.argv) > 2:
        to_excel(sys.argv[2], date, df)


if __name__ == "__main__":
    main()
