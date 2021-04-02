#! /usr/bin/python3
from datetime import datetime
from pprint import pprint
import sys

import openpyxl
from jqdatasdk import *
import pandas

# A股代号: (中文名称, 未来3年净利润增长率, 理想市盈率)
# 未来3年净利润增长率和理想市盈率来自以前的好公司分析
# 好价格 = 当期净利润 * (1 + 未来3年净利润增长率) ^ 3 * 合理市盈率 / 总股本 / 2
# 合理价格 = 当期净利润 * (1 + 未来3年净利润增长率) * 合理市盈率 / 总股本

Stocks = {
    '000002.XSHE': ('万科A',  0.15, 10),
    '000333.XSHE': ('美的集团', 0.10, 15),
    '000651.XSHE': ('格力电器', 0.10, 12),
    '000858.XSHE': ('五粮液',  0.15, 25),
    '000895.XSHE': ('双汇发展', 0.06, 19),
    '002714.XSHE': ('牧原股份', 0.40, 15),
    '002271.XSHE': ('东方雨虹', 0.30, 15),
    '002304.XSHE': ('洋河股份', 0.08, 20),
    '002372.XSHE': ('伟星新材', 0.09, 20),
    '002415.XSHE': ('海康威视', 0.10, 24),
    '002508.XSHE': ('老板电器', 0.10, 17),
    '002677.XSHE': ('浙江美大', 0.15, 15),
    '300015.XSHE': ('爱尔眼科', 0.30, 60),
    '600009.XSHG': ('上海机场', 0.05, 22),
    '600104.XSHG': ('上汽集团', 0.00, 8),
    '600276.XSHG': ('恒瑞医药', 0.20, 60),
    '600309.XSHG': ('万华化学', 0.58, 15),
    '600519.XSHG': ('贵州茅台', 0.15, 30),
    '600585.XSHG': ('海螺水泥', 0.15, 8),
    '600660.XSHG': ('福耀玻璃', 0.02, 16),
    '600887.XSHG': ('伊利股份', 0.10, 23),
    '600900.XSHG': ('长江电力', 0.08, 20),
    '601318.XSHG': ('中国平安', 0.20, 9),
    '601668.XSHG': ('中国建筑', 0.10, 7),
    '603288.XSHG': ('海天味业', 0.15, 25),
    '603886.XSHG': ('元祖股份', 0.10, 15)
}

Labels = ['Code', '中文名称', '2019归母净利润', '总股本', '未来3年净利润增长率', '理想市盈率', '好价格', '合理价格', '收盘',
          '收盘/合理价格']


def save_to_spreadsheet(filename: str, sheet_name: str, result: list):
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
    except KeyError:
        sheet = wb.create_sheet(title=sheet_name)

    row = sheet.max_row
    if row == 1:
        for j in range(len(Labels)):
            c = sheet.cell(row=1, column=j+1)
            c.value = Labels[j]

    for i in range(len(result)):
        for key in result[i].keys():
            try:
                j = Labels.index(key)
                c = sheet.cell(row=row+i+1, column=j+1)
                if key in [Labels[0], Labels[1]]:
                    pass
                elif key in [Labels[3], Labels[5]]:
                    c.number_format = '#,##,0'
                elif key == Labels[4]:
                    c.number_format = '0%'
                else:
                    c.number_format = '#,##,0.00'
                c.value = result[i][key]
            except ValueError:
                pass
    wb.save(filename)


def good_prices(codes: list, stat_date: str, date: str) -> list:
    df = get_price(codes, count=1, end_date=date, frequency='daily', fields=['close'], panel=False)
    df2 = get_fundamentals(query(valuation).filter(indicator.code.in_(codes)), statDate=stat_date)
    df3 = get_fundamentals(query(income).filter(income.code.in_(codes)), statDate=stat_date)
    df.index = df['code']
    df2.index = df2['code']
    df3.index = df3['code']
    df['capitalization'] = df2['capitalization']
    df['np_parent_company_owners'] = df3['np_parent_company_owners']
    # print(len(df))
    # print(df)
    result = []
    for i in range(len(df)):
        code = df['code'][i]
        profit = df['np_parent_company_owners'][i]
        cap = df['capitalization'][i] * 10000
        close = df['close'][i]
        name = Stocks[code][0]
        inc = Stocks[code][1]
        pe = Stocks[code][2]
        good = profit * ((1 + inc) ** 3) * pe / cap / 2
        reasonable = profit * (1 + inc) * pe / cap
        dic = {Labels[0]: ('SZ:' if code.endswith('.XSHE') else 'SH:') + code[: 6],
               Labels[1]: name,
               Labels[2]: profit,
               Labels[3]: cap,
               Labels[4]: inc,
               Labels[5]: pe,
               Labels[6]: good,
               Labels[7]: reasonable,
               Labels[8]: close,
               Labels[9]: close / reasonable}
        result.append(dic.copy())
    return sorted(result, key=lambda x: x[Labels[9]])


def main():
    if len(sys.argv) < 4:
        print('Usage: {} stat_date(%Y) date(%Y-%m-%d) xlsx'.format(sys.argv[0]))
        sys.exit(1)

    with open('auth/jq_key.txt', 'r') as f:
        account = f.readline()[: -1]
        password = f.readline()[: -1]
    auth(account, password)

    pandas.set_option('display.max_rows', 100)
    pandas.set_option('display.max_columns', 50)

    result = good_prices(list(Stocks.keys()), sys.argv[1], sys.argv[2])
    pprint(result)
    save_to_spreadsheet(sys.argv[3], sys.argv[2], result)


if __name__ == "__main__":
    main()
