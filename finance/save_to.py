#! /usr/bin/python3
import sys
from datetime import datetime

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter
from pymongo import MongoClient


def save_to_spreadsheet(filename: str, sheet_name: str, result: list):
    titles = ['platform', 'currency', 'code', 'name', 'risk', 'market_value', 'hold_gain',
              'mv_rmb', 'hg_rmb']

    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheet_name]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
    except KeyError:
        wb.create_sheet(title=sheet_name)

    row = sheet.max_row
    if row == 1:
        for j in range(len(titles)):
            c = sheet.cell(row=1, column=j+1)
            c.value = titles[j]

    for i in range(len(result)):
        for key in result[i].keys():
            try:
                j = titles.index(key)
                c = sheet.cell(row=row+i+2, column=j+1)
                if key in ['market_value', 'hold_gain', 'mv_rmb', 'hg_rmb']:
                    c.number_format = "#,##,0.00"
                    c.value = result[i][key]
                else:
                    c.value = result[i][key]
            except ValueError:
                pass
        if 'currency' not in result[i].keys():
            j = titles.index('currency')
            sheet.cell(row=row+i+2, column=j+1).value = 'rmb'
        if 'mv_rmb' not in result[i].keys():
            j = titles.index('mv_rmb')
            c = sheet.cell(row=row+i+2, column=j+1)
            c.number_format = "#,##,0.00"
            c.value = result[i]['market_value']
        if 'hg_rmb' not in result[i].keys():
            j = titles.index('hg_rmb')
            c = sheet.cell(row=row+i+2, column=j+1)
            c.number_format = "#,##,0.00"
            c.value = result[i]['hold_gain']
    wb.save(filename)


def save_to_mongo(collection: str, result: list):
    mongo_host = '127.0.0.1'
    mongo_port = 27017
    mongo_db_name = 'finance'

    client = MongoClient(host=mongo_host, port=mongo_port)
    db = client[mongo_db_name]
    collection = db[collection]
    collection.insert_many(result)


def summarize_amount(file: str, sheet_name: str):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheet_name]
    last_row = sheet.max_row

    summaries = [
        {'location': (last_row + 2, 1),
         'letter': 'A',
         'labels': ['银河', '华盛HKD', '华盛USD', '蛋卷*', '同花顺'],
         'category': 'platform',
         'anchor': 'K1'
         },
        {'location': (last_row + 2, 4),
         'letter': 'B',
         'labels': ['rmb', 'hkd', 'usd'],
         'category': 'currency',
         'anchor': 'K16'
         },
        {'location': (last_row + 2, 7),
         'letter': 'E',
         'labels': [0, 1, 2, 3],
         'category': 'risk',
         'anchor': 'K31'
         }
    ]

    for summary in summaries:
        row, col = summary['location']
        le = [get_column_letter(j) for j in range(col, col + 3)]
        for i in range(len(summary['labels'])):
            sheet.cell(row=row+i, column=col).value = summary['labels'][i]
            c = sheet.cell(row=row+i, column=col+1)
            c.number_format = "#,##,0.00"
            c.value = '=SUMIF(${0}$2:${0}${1},{2}{3},$H$2:$H${1})'.format(summary['letter'],
                                                                          last_row, le[0], row + i)
            c = sheet.cell(row=row+i, column=col+2)
            c.number_format = "#,##,0.00"
            c.value = '=SUMIF(${0}$2:${0}${1},{2}{3},$I$2:$I${1})'.format(summary['letter'],
                                                                          last_row, le[0], row + i)
        sheet.cell(row=row+i+1, column=col).value = 'sum'
        c = sheet.cell(row=row+i+1, column=col+1)
        c.number_format = "#,##,0.00"
        c.value = '=SUM({0}{1}:{0}{2})'.format(le[1], row, row + i)
        c = sheet.cell(row=row+i+1, column=col+2)
        c.number_format = "#,##,0.00"
        c.value = '=SUM({0}{1}:{0}{2})'.format(le[2], row, row + i)

        pie = PieChart()
        labels = Reference(sheet, min_col=col, min_row=row, max_row=row+i)
        data = Reference(sheet, min_col=col+1, min_row=row-1, max_row=row+i)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = summary['category']
        sheet.add_chart(pie, summary['anchor'])

    wb.save(file)


def main():
    if len(sys.argv) < 3:
        print('Usage: {} file sheet'.format(sys.argv[0]))
        sys.exit(1)

    summarize_amount(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
