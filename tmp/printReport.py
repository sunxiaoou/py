#! /usr/local/bin/python3

from __future__ import annotations

import glob
import os
import re
from typing import Optional, Tuple

import openpyxl


def _find_today_progress_header(ws) -> Optional[Tuple[int, int, int]]:
    """
    返回 (header_row, col_work, col_status)
    header_row: '工作内容'/'完成情况' 所在行
    col_work/col_status: 列号(1-based)
    """
    max_r, max_c = ws.max_row, ws.max_column

    # 先找包含“今日进展”的行
    today_row = None
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "今日进展" in v:
                today_row = r
                break
        if today_row:
            break

    if not today_row:
        return None

    # 标题一般就在同一行（示例里是：A2=今日进展，B2=工作内容，C2=完成情况）
    header_row = today_row
    col_work = col_status = None

    for c in range(1, max_c + 1):
        v = ws.cell(header_row, c).value
        if v == "工作内容":
            col_work = c
        elif v == "完成情况":
            col_status = c

    # 兜底：如果同一行没找到，尝试往下 1~3 行找标题
    if col_work is None or col_status is None:
        for rr in range(today_row, min(today_row + 3, max_r) + 1):
            cw = cs = None
            for c in range(1, max_c + 1):
                v = ws.cell(rr, c).value
                if v == "工作内容":
                    cw = c
                elif v == "完成情况":
                    cs = c
            if cw and cs:
                header_row, col_work, col_status = rr, cw, cs
                break

    if col_work is None or col_status is None:
        return None

    return header_row, col_work, col_status


def print_today_progress_from_sheet(ws, sheet_label: str, file_label: str) -> None:
    found = _find_today_progress_header(ws)
    if not found:
        return

    header_row, col_work, col_status = found

    print(f"\n=== {os.path.basename(file_label)} | Sheet: {sheet_label} ===")

    blank_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        # 遇到“明日计划”就停止
        row_has_tomorrow = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "明日计划" in v:
                row_has_tomorrow = True
                break
        if row_has_tomorrow:
            break

        work = ws.cell(r, col_work).value
        status = ws.cell(r, col_status).value

        # 全空行：累计，连续空行太多就结束（防止拖到表尾）
        if (work is None or str(work).strip() == "") and (status is None or str(status).strip() == ""):
            blank_streak += 1
            if blank_streak >= 3:
                break
            continue
        blank_streak = 0

        # 只要“工作内容”非空就打印（示例里有“项目大标题”那种行，完成情况为空也会打印）
        if work is not None and str(work).strip() != "":
            w = str(work).strip()
            s = "" if status is None else str(status).strip()
            print(f"- {w} | {s}")


def is_workday_sheet(name: str) -> bool:
    # 过滤掉类似 WpsReserved_* 这种
    if name.lower().startswith("wpsreserved"):
        return False
    # 你的示例是 250108 这种纯数字日期；这里放宽一点：只要包含数字就当候选
    return bool(re.fullmatch(r"\d{6,8}", name)) or any(ch.isdigit() for ch in name)


def main():
    # 当前目录下所有xlsx；如需指定目录，改成比如 "reports/*.xlsx"
    files = sorted(glob.glob("report/孙笑鸥-日报*.xlsx"))
    if not files:
        print("当前目录未找到 .xlsx 文件")
        return

    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True)
        for sh in wb.sheetnames:
            if not is_workday_sheet(sh):
                continue
            ws = wb[sh]
            print_today_progress_from_sheet(ws, sheet_label=sh, file_label=f)


if __name__ == "__main__":
    main()
