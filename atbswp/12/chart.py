#! /usr/bin/python3
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series, PieChart, ScatterChart
from openpyxl.chart.layout import ManualLayout, Layout
from openpyxl.chart.marker import DataPoint


def bar_chart():
    wb = Workbook()
    sheet = wb.active
    for i in range(1, 11): # create some data in column A
        sheet['A' + str(i)] = i
    refObj = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
    seriesObj = Series(refObj, title='First series')
    chartObj = BarChart()
    chartObj.title = 'My Chart'
    chartObj.append(seriesObj)
    sheet.add_chart(chartObj, 'C5')
    wb.save('chart.xlsx')


def pie_chart():
    data = [
        ['Pie', 'Sold'],
        ['Apple', 50],
        ['Cherry', 30],
        ['Pumpkin', 10],
        ['Chocolate', 40],
    ]
    wb = Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Pies sold by category"

    # Cut the first slice out of the pie
    slice = DataPoint(idx=0, explosion=20)
    pie.series[0].data_points = [slice]

    ws.add_chart(pie, "D1")
    wb.save("chart.xlsx")


def layout():
    wb = Workbook()
    ws = wb.active
    rows = [
        ['Size', 'Batch 1', 'Batch 2'],
        [2, 40, 30],
        [3, 40, 25],
        [4, 50, 30],
        [5, 30, 25],
        [6, 25, 35],
        [7, 20, 40],
    ]

    for row in rows:
        ws.append(row)
    ch1 = ScatterChart()
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
    for i in range(2, 4):
        values = Reference(ws, min_col=i, min_row=1, max_row=7)
        series = Series(values, xvalues, title_from_data=True)
        ch1.series.append(series)

    ch1.title = "Default layout"
    ch1.style = 13
    ch1.x_axis.title = 'Size'
    ch1.y_axis.title = 'Percentage'
    ch1.legend.position = 'r'

    ws.add_chart(ch1, "B10")

    from copy import deepcopy

    # Half-size chart, bottom right
    ch2 = deepcopy(ch1)
    ch2.title = "Manual chart layout"
    ch2.legend.position = "tr"
    ch2.layout=Layout(
        manualLayout=ManualLayout(
            x=0.25, y=0.25,
            h=0.5, w=0.5,
        )
    )
    ws.add_chart(ch2, "H10")

    # Half-size chart, centred
    ch3 = deepcopy(ch1)
    ch3.layout = Layout(
        ManualLayout(
            x=0.25, y=0.25,
            h=0.5, w=0.5,
            xMode="edge",
            yMode="edge",
        )
    )
    ch3.title = "Manual chart layout, edge mode"
    ws.add_chart(ch3, "B27")

    # Manually position the legend bottom left
    ch4 = deepcopy(ch1)
    ch4.title = "Manual legend layout"
    ch4.legend.layout = Layout(
        manualLayout=ManualLayout(
            yMode='edge',
            xMode='edge',
            x=0, y=0.9,
            h=0.1, w=0.5
        )
    )

    ws.add_chart(ch4, "H27")
    wb.save("chart.xlsx")


def main():
    # bar_chart()
    pie_chart()
    # layout()


if __name__ == "__main__":
    main()
