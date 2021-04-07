#! /usr/bin/python3
import datetime
import re

from openpyxl import Workbook


def main():
    wb = Workbook()
    ws = wb.active
    # set date using a Python datetime
    ws['A1'] = datetime.datetime(2010, 7, 21)
    print(ws['A1'].number_format)
    # ws['A2'] = float(re.sub(",", "", "34,175,619.57"))

    cell = ws.cell(row=2, column=1)
    cell.number_format = "#,##,0.00"
    cell.value = float(re.sub(",", "", "34,175,619.57"))
    # print(ws['A2'].number_format)
    cell = ws.cell(row=3, column=1)
    # cell.number_format = '0.00%'
    cell.style = "Percent"
    cell.number_format = "0.00%"
    cell.value = 0.5382
    # cell.value = "2.53%"
    wb.save("numformat.xlsx")


if __name__ == "__main__":
    main()
