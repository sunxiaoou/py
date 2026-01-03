#! /usr/bin/python3
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def autosize_and_freeze(writer, sheet_name: str, df: pd.DataFrame):
    """Auto column width + freeze header row."""
    ws = writer.sheets[sheet_name]
    ws.freeze_panes = "A2"
    # Auto width (simple heuristic)
    for i, col in enumerate(df.columns, start=1):
        series = df[col].astype(str)
        max_len = max([len(col)] + series.map(len).tolist()) if len(series) else len(col)
        max_len = min(max_len + 2, 60)  # cap
        ws.column_dimensions[get_column_letter(i)].width = max_len

def df_to_sheet(df: pd.DataFrame, xlsx: str, sheet: str, overlay=False, header=True):
    if not os.path.exists(xlsx):
        with pd.ExcelWriter(xlsx, engine='openpyxl') as w:
            df.to_excel(w, sheet_name=sheet, index=False, header=header)
            autosize_and_freeze(w, sheet, df)
        print(f"Created new Excel file with sheet({sheet})")
    elif not overlay:
        with pd.ExcelWriter(xlsx, engine='openpyxl', mode='a') as w:
            if sheet in w.book.sheetnames:
                print(f"Sheet({sheet}) already exists in {xlsx}")
                return
            df.to_excel(w, sheet_name=sheet, index=False, header=header)
            autosize_and_freeze(w, sheet, df)
        print(f"Added new sheet({sheet}) to file {xlsx}")
    else:
        with pd.ExcelWriter(xlsx, engine='openpyxl', mode='a', if_sheet_exists='overlay') as w:
            df.to_excel(w, sheet_name=sheet, index=False, header=header, startrow=0, startcol=0)
            autosize_and_freeze(w, sheet, df)
        print(f"Overlap sheet({sheet}) in file {xlsx}")


def duplicate_last_sheet(xlsx: str, sheet: str, row_count: int):
    wb = load_workbook(xlsx)
    last_sheet = wb.worksheets[-1]
    new_sheet = wb.copy_worksheet(last_sheet)
    if row_count < new_sheet.max_row:
        new_sheet.delete_rows(row_count, new_sheet.max_row - 1)
    new_sheet.title = sheet
    wb.active = wb.index(new_sheet)
    wb.save(xlsx)


def main():
    pass


if __name__ == "__main__":
    main()
