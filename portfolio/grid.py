#! /usr/bin/python3
import re
import sys
import time

import numpy as np
import pandas as pd
from matplotlib import dates, pyplot, ticker
from openpyxl import load_workbook

from mysql import MySql
from xueqiu import Xueqiu

# pd.set_option('display.max_rows', 4000)
pd.set_option('display.max_columns', 10)


class Grid:
    def __init__(self, low: float, high: float, number: int, is_percent: bool):
        self.low = low
        self.high = high
        self.is_percent = is_percent
        self.number = number
        if not self.is_percent:
            self.change = round((high - low) / number, 2)
            self.change2 = - self.change
            self.array = [high - self.change * i for i in range(self.number + 1)]
        else:
            self.change = round((high / low) ** (1 / number) - 1, 4)
            self.change2 = round((low / high) ** (1 / number) - 1, 4)
            self.array = [round(high / (1 + self.change) ** i, 2) for i in range(self.number + 1)]

    def __str__(self):
        if self.is_percent:
            changes = '({}%, {}%)'.format(round(self.change * 100, 2), round(self.change2 * 100, 2))
        else:
            changes = '(%.2f, %.2f)' % (self.change, self.change2)
        return self.to_str() + ': ' + str(changes) + ' ' + str(self.array)

    def get_count(self, price: float, benchmark: float = -1) -> tuple:
        index = 0 if benchmark == -1 else self.array.index(benchmark)
        count = 0
        while index < self.number and self.array[index + 1] >= price:
            index += 1
            count += 1
        if count:
            return index, np.nan if index > len(self.array) - 2 else self.array[index + 1], self.array[index], count

        while index > 0 and self.array[index - 1] <= price:
            index -= 1
            count -= 1
        return index, np.nan if index > len(self.array) - 2 else self.array[index + 1], self.array[index], count

    def show_grid(self, quantity: int) -> pd.DataFrame:
        df = pd.DataFrame([self.array, [0] + [quantity] * self.number],
                          index=['price', 'quantity']).T
        df['quantity'] = df['quantity'].apply(lambda x: int(x))
        df['q_sum'] = df['quantity'].cumsum()
        df['amount'] = df['price'] * df['quantity']
        df['cost'] = df['amount'].cumsum()
        df['value'] = df['price'] * df['q_sum']
        df['loss'] = df['value'] - df['cost']
        df['loss(%)'] = round(df['loss'] / df['cost'] * 100, 2)
        return df

    def to_str(self) -> str:
        return "%d_%d_%d_%d" % (self.low, self.high, self.number, 1 if self.is_percent else 0)

    @classmethod
    def make(cls, s: str):
        args = [int(i) for i in s.split('_')]
        args[-1] = True if args[-1] else False
        return cls(*args)


class LoopBack:
    @staticmethod
    def complete_code(code: str) -> str:
        if len(code) == 8 and code[: 2] in ['sh', 'sz', 'SH', 'SZ'] and code[2:].isnumeric():
            code = code.upper()
        elif len(code) == 6 and code.isnumeric():
            if code.startswith('11'):
                code = 'SH' + code
            elif code.startswith('12'):
                code = 'SZ' + code
        return code

    def __init__(self, grid: Grid, code: str, quantity: int, start_date: str = ''):
        code = LoopBack.complete_code(code)
        db = MySql(database='portfolio')
        where = 'code = "%s"' % code
        if start_date:
            where += ' AND date >= "%s"' % start_date
        data = db.to_frame('cvtbone_daily', ['date', 'name', 'open'], where)
        if data.empty:
            snowball = Xueqiu()
            data = snowball.get_data(code, start_date)
            data['date'] = data['date'].apply(lambda x: x.date())
            data['code'] = code
            data['name'] = snowball.get_name(code)
            data = data[['date', 'name', 'open']]
        name = data['name'].iloc[0]

        self.grid = grid
        self.code = code
        self.name = name
        self.data = data
        self.quantity = quantity
        self.benchmark = grid.high
        self.index = 0
        self.cost = 0
        self.value = 0
        self.trans = []

    def __str__(self):
        return 'index(%d), benchmark(%.2f), value(%.2f), cost(%.2f), profit(%.2f)' %\
               (self.index, self.benchmark, self.value, self.cost, self.value - self.cost)

    def trade_open(self, day: str, price: float):
        self.index, _, self.benchmark, count = self.grid.get_count(price, self.benchmark)
        self.value = price * self.quantity * self.index
        if count:
            volume = self.quantity * count
            self.cost += price * volume
            dic = {'date': day,
                   'price': price,
                   'volume': volume,
                   'index': self.index,
                   'value': self.value,
                   'cost': self.cost,
                   'profit': round(self.value - self.cost, 2)}
            self.trans.append(dic.copy())

    def draw(self, data: pd.DataFrame, trans: pd.DataFrame):
        data = data[['date', 'open']]
        data = data.rename({'open': self.code}, axis=1)
        data = data.dropna().set_index('date')
        data.index.name = None
        # print(data)

        buy = trans[trans['volume'] > 0][['date', 'price']]
        buy = buy.dropna().set_index('date', drop=True)
        buy.index.name = None
        sell = trans[trans['volume'] < 0][['date', 'price']]
        sell = sell.dropna().set_index('date', drop=True)
        sell.index.name = None

        pyplot.figure(figsize=(10, 6))
        pyplot.title('网格策略回测 %s(%s) (%s ~ %s)' %
                     (self.name, self.code, data.index[0], data.index[-1]))
        pyplot.ylabel('开盘价(元)')
        pyplot.grid()
        pyplot.gca().xaxis.set_major_locator(ticker.MultipleLocator(data.shape[0] // 8))
        pyplot.gca().xaxis.set_major_formatter(dates.DateFormatter('%y-%m-%d'))
        pyplot.xticks(rotation=30)
        pyplot.plot(data.index, data[self.code])
        pyplot.plot(buy.index, buy['price'], 'og')
        pyplot.plot(sell.index, sell['price'], 'or')

        pyplot.figtext(0.9, 0.85, ' Grid')
        pyplot.figtext(0.9, 0.80, ' 格数 %d' % (self.grid.number + 1))
        pyplot.figtext(0.9, 0.76, ' 最高 %.2f' % self.grid.high)
        pyplot.figtext(0.9, 0.72, ' 最低 %.2f' % self.grid.low)
        if not self.grid.is_percent:
            pyplot.figtext(0.9, 0.68, ' 涨幅 %.2f' % self.grid.change)
            pyplot.figtext(0.9, 0.64, ' 跌幅 %.2f' % self.grid.change2)
        else:
            pyplot.figtext(0.9, 0.68, ' 涨幅 %.2f' % (self.grid.change * 100) + '%')
            pyplot.figtext(0.9, 0.64, ' 跌幅 %.2f' % (self.grid.change2 * 100) + '%')
        pyplot.figtext(0.9, 0.60, ' 数量 %d' % self.quantity)
        pyplot.figtext(0.9, 0.55, ' Trade')
        pyplot.figtext(0.9, 0.50, ' 开盘买(绿)')
        pyplot.figtext(0.9, 0.46, ' 开盘卖(红)')
        pyplot.figtext(0.9, 0.41, ' Result')
        pyplot.figtext(0.9, 0.36, ' 剩余网格 %d' % self.index)
        pyplot.figtext(0.9, 0.32, ' 基准 %.2f' % self.benchmark)
        pyplot.figtext(0.9, 0.28, ' 市值 %.2f' % self.value)
        pyplot.figtext(0.9, 0.24, ' 成本 %.2f' % self.cost)
        pyplot.figtext(0.9, 0.20, ' 盈亏 %.2f' % (self.value - self.cost))
        pyplot.figtext(0.9, 0.12, ' - 同光和尘')
        pyplot.savefig('%s_%s.png' % (self.code, self.name), dpi=400, bbox_inches='tight')
        pyplot.show()

    def trade_daily(self) -> str:
        for _, row in self.data.iterrows():
            # print(row['date'], row['open'])
            self.trade_open(row['date'], row['open'])

        if 'pic' not in sys.argv:
            return '%s(%s) %s_%s %d %d %.2f %.2f %.2f %.2f' \
                   % (self.code, self.name, self.data.iloc[0]['date'], self.data.iloc[-1]['date'], len(self.trans),
                      self.index * self.quantity, self.benchmark, self.value, self.cost, self.value - self.cost)
        else:
            trans = pd.DataFrame(self.trans)
            print(trans)
            print('%s~%s' % (self.data.iloc[0]['date'], self.data.iloc[-1]['date']), self)
            self.draw(self.data, trans)
            return ''


GRID_ARGS = [
    (105, 115, 5, False),
    (105, 115, 5, True),
    (110, 125, 5, False),
    (110, 125, 5, True),
    (115, 135, 5, False),
    (115, 135, 5, True),
    (120, 150, 5, False),
    (120, 150, 5, True),
    (125, 165, 5, False),
    (125, 165, 5, True)
]


def get_codes(file: str) -> list:
    with open(file) as f:
        text = f.read()
    blocks = text.split('代码')
    lines1 = blocks[1].split('\n')
    l1 = [row.split()[1] for row in lines1[1: -2]]
    lines5 = blocks[5].split('\n')
    l5 = [row.split()[1] for row in lines5[1: -2]]
    lines6 = blocks[6].split('\n')
    l6 = [row.split()[1] for row in lines6[1: -2]]
    if len(blocks) > 7:
        lines7 = blocks[7].split('\n')
        l7 = [row.split()[1] for row in lines7[1: -1]]
    else:
        l7 = []
    return sorted(list(set(l1 + l5 + l6 + l7)))


def trade_codes(grid: Grid, codes: list, quantity: int, start_date: str) -> pd.DataFrame:
    result = []
    for code in codes:
        s = LoopBack(grid, code, quantity, start_date).trade_daily()
        s = s.split()
        result.append([s[0], float(s[-1])])
        # time.sleep(0.1)
    return pd.DataFrame(result, columns=['code_name', grid.to_str()])


def get_count(row: pd.Series) -> tuple:
    _, bm, bm2, count = Grid.make(row['max_col_name']).get_count(row['price'])
    if bm2 < row['price']:
        bm, bm2 = bm2, np.nan
    return bm, bm2, count


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError:
        df.to_excel(xlsx, sheet_name=sheet, index=False)
        print(xlsx + ' created')
        return

    ws = wb.copy_worksheet(wb.worksheets[-1])       # copy a old sheet as template to avoid adjust size
    if df.shape[0] < ws.max_row:
        ws.delete_rows(df.shape[0], ws.max_row - 1)
    ws.title = sheet
    wb.active = len(wb.worksheets) - 1

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()


def batch(file: str, quantity: int, start_date: str):
    codes = [LoopBack.complete_code(i) for i in get_codes(file)]
    snowball = Xueqiu()
    dic = snowball.last_close(codes[0])
    date = dic['date'].strftime('%y%m%d')
    df = snowball.get_cvtbones(codes)
    result = df[['code', 'price']]

    for i, args in enumerate(GRID_ARGS):
        df = trade_codes(Grid(*args), codes, quantity, start_date)
        if i == 0:
            df['code'] = df.apply(lambda x: x['code_name'][: 8], axis=1)
            result = result.merge(df, on='code', how='inner')
        else:
            result = result.merge(df, on='code_name', how='inner')

    n = len(GRID_ARGS)
    result['max_col_name'] = result.iloc[:, -n:].idxmax(axis=1)
    result['max_value'] = result[result.columns[-n - 1: -1]].max(axis=1)
    result['BM'] = result.apply(lambda x: get_count(x)[0], axis=1)
    result = result[['code_name', 'max_value', 'max_col_name', 'BM', 'price']]
    result['BM2'] = result.apply(lambda x: get_count(x)[1], axis=1)
    result['count'] = result.apply(lambda x: quantity * get_count(x)[2], axis=1)

    result = result.sort_values(by='max_value', ascending=False)
    result.reset_index(drop=True, inplace=True)
    result = result.reset_index()       # convert index to column
    # with open('grid.html', 'w') as f:
    #     f.write(df.to_html())
    result['amount'] = result.apply(lambda x: '=F%d*H%d' % (x['index'] + 2, x['index'] + 2), axis=1)
    # print(df)
    to_excel('grid.xlsx', date, result)


def show_grids(quantity: int):
    for args in GRID_ARGS:
        grid = Grid(*args)
        print(grid)
        if quantity:
            print(grid.show_grid(quantity))


def usage():
    print('Usage: {} grid low,high,number,is_percent price benchmark [quantity]'.
          format(sys.argv[0]))
    print('       %s grid [quantity]' % sys.argv[0])
    print('       {} loopback low,high,number,is_percent code quantity [pic start_date(%Y-%m-%d)]'.
          format(sys.argv[0]))
    print('       {} batch cvt_file quantity start_date(%Y-%m-%d)'.format(sys.argv[0]))


def main():
    if len(sys.argv) > 4:
        if sys.argv[1] == 'grid':
            a = sys.argv[2].split(',')
            grid = Grid(float(a[0]), float(a[1]), int(a[2]), True if int(a[3]) else False)
            print(grid)
            if len(sys.argv) > 5:
                print(grid.show_grid(int(sys.argv[5])))
            print(grid.get_count(float(sys.argv[3]), float(sys.argv[4])))
        elif sys.argv[1] == 'loopback':
            a = sys.argv[2].split(',')
            grid = Grid(float(a[0]), float(a[1]), int(a[2]), True if int(a[3]) else False)
            if re.match(r'\d\d\d\d-\d\d-\d\d', sys.argv[-1]):
                loopback = LoopBack(grid, sys.argv[3], int(sys.argv[4]), sys.argv[-1])
            else:
                loopback = LoopBack(grid, sys.argv[3], int(sys.argv[4]))
            print(loopback.trade_daily())
        elif sys.argv[1] == 'batch':
            batch(sys.argv[2], int(sys.argv[3]), sys.argv[4])
        else:
            usage()
    elif len(sys.argv) > 1 and sys.argv[1] == 'grid':
        if len(sys.argv) > 2:
            show_grids(int(sys.argv[2]))
        else:
            show_grids(0)
    else:
        usage()


if __name__ == '__main__':
    main()
