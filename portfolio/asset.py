#! /usr/bin/python3
import os
import re
import sys
from datetime import datetime
from pprint import pprint

import numpy as np
import openpyxl
import pandas as pd
import pyperclip
import requests
from bs4 import BeautifulSoup
from forex_python.converter import CurrencyRates
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

from mysql import MySql
from securities import *

pd.set_option('display.max_rows', 200)
# pd.set_option('display.max_columns', 10)


COLUMNS = ['platform', 'currency', 'code', 'name', 'type', 'risk', 'market_value', 'hold_gain']


def verify(row: pd.Series):
    if not pd.isna(row['nav']):
        mv = round(row['nav'] * row['volume'], 2)
        if row['market_value'] != mv:
            print(row['nav'], type(row['nav']))
            print('Warning: {} market_value({}) != {}'.format(row['name'], row['market_value'], mv))
        hg = round((row['nav'] - row['cost']) * row['volume'], 2)
        if abs(row['hold_gain'] - hg) >= 1:
            print('Warning: {} hold_gain({}) != {}'.format(row['name'], row['hold_gain'], hg))


def zhaoshang_bank(datafile: str) -> pd.DataFrame:
    with open(datafile) as f:
        lines = []
        for line in f.readlines():
            lines += re.sub(r'[,，:>]', '', line).rstrip('\n').split()

    i = 0
    while lines[i] != '尾号8884':
        i += 1
    cash = float(lines[i + 1])
    result = [('招商银行', 'cny', 'cash', '现金', '货币', 0, cash, 0)]
    i += 2
    while not lines[i].startswith('理财'):
        i += 1
    i += 1
    while not re.match(r'^[.\d]+$', lines[i]):
        i += 1
    total_mv = float(lines[i])
    asset = round(cash + total_mv, 2)
    i += 1
    try:
        while lines[i]:
            code = lines[i][4:] if lines[i].startswith('招银理财') else lines[i]
            if code == '保险':
                break
            if code.startswith('招赢日日盈'):
                code = '招赢日日盈'
            i += 1
            while not re.match(r'.*[\d.]+', lines[i]):
                i += 1
            hold_gain = float(re.sub('[^\d.]+', '', lines[i]))
            i += 1
            while not re.match(r'.*[\d.]+', lines[i]):
                i += 1
            market_value = float(lines[i])
            name, type, risk = SECURITIES[code]
            result.append(('招商银行', 'cny', code, name, type, risk, market_value, hold_gain))
            i += 4
    except IndexError:
        pass
    df = pd.DataFrame(result, columns=COLUMNS)
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    return df


def hangseng_bank(datafile: str) -> pd.DataFrame:
    cash = float(re.sub(r'.+_', '', datafile[: -4]))
    result = [('恒生银行', 'cny', 'cash', '现金', '货币', 0, cash, 0)]

    with open(datafile) as fp:
        lines = [re.sub(r'[,＋]', '', re.sub('－', '-', line)).rstrip('\n') for line in fp.readlines()]
    i = 0
    while not lines[i].startswith('总市值'):
        i += 1
    total_mv = float(lines[i + 1])
    asset = round(cash + total_mv, 2)
    total_hg = float(re.sub(r'[^-\d.]+', '', lines[i + 2]))
    i += 3
    # print(cash, total_mv, total_hg)
    while lines[i] != '交易记录':
        i += 1
    i += 1
    while i < len(lines) and re.match(r'\d{6}.*', lines[i]):
        if len(lines[i]) == 6:
            code = 'F' + lines[i]
            i += 1
        else:
            code = 'F' + lines[i][: 6]
        name, type, risk = SECURITIES[code]
        i += 1
        if not re.match(r'.*[\d.]+', lines[i]):
            i += 1
        hold_gain = float(lines[i])
        i += 6
        if re.match(r'.*[\d.]+', lines[i]):
            market_value = float(re.sub(r'[^\d.]+', '', lines[i]))
        else:
            i += 1
            market_value = float(lines[i])
        result.append(('恒生银行', 'cny', code, name, type, risk, market_value, hold_gain))
        i += 1
    df = pd.DataFrame(result, columns=COLUMNS)
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    return df


def yinhe(datafile: str) -> pd.DataFrame:
    s = re.sub(r'.+_', '', datafile[: -4])
    cash2 = float(s) if s else 0

    with open(datafile) as f:
        lines = []
        for line in f.readlines():
            lines += re.sub('－', '-', line).rstrip('\n').split()
    i = 0
    while not lines[i].startswith('场内资产'):
        i += 1
    i += 1
    while not re.match(r'^[.\d]+$', lines[i]):
        i += 1
    asset = float(lines[i])
    i += 1
    while not re.match(r'^[.\d]+$', lines[i]):
        i += 1
    total_mv = float(lines[i])
    i += 1
    while not re.match(r'^[-.\d]+$', lines[i]):
        i += 1
    total_hg = float(lines[i])
    i += 1
    while not re.match(r'^[.\d]+$', lines[i]):
        i += 1
    cash = float(lines[i])
    i += 1
    assert round(total_mv + cash, 2) == asset, \
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('银河', 'cny', 'cash', '现金', '货币', 0, cash + cash2, 0)]
    i += 8
    while not lines[i].startswith('参考盈亏'):
        i += 1
    i += 1
    while True:
        while i < len(lines) and not re.match(r".*\d{6}$", lines[i]):
            i += 1
        if i == len(lines):
            break
        if len(lines[i]) == 6:
            name, code = lines[i - 1], lines[i]
        else:
            name, code = lines[i][: -6].rstrip(), lines[i][-6:]
        i += 1
        if code[0] in ['1', '7']:
            type, risk = '转债', 2
        elif code[0] == '2':
            type, risk = '货币', 0
        else:
            name, type, risk = SECURITIES[code]

        hold_gain = float(lines[i])
        i += 1
        if ' ' in lines[i]:  # volume, cost are in same line
            l = lines[i].split()
            volume, cost = int(l[0]), float(l[1])
            i += 1
        else:
            volume, cost = int(lines[i]), float(lines[i + 1])
            i += 2
        market_value = float(lines[i])
        i += 2
        if ' ' in lines[i]:
            l = lines[i].split()
            v2, nav = int(l[0]), float(l[1])
            i += 1
        else:
            v2, nav = int(lines[i]), float(lines[i + 1])
            i += 2
        assert v2 == volume, print("v2{} != volume{}".format(v2, volume))
        assert market_value == round(nav * volume, 2), \
            print("mv({}) != nav({}) * volume({})".format(market_value, nav, volume))
        result.append(('银河', 'cny', code, name, type, risk, market_value, hold_gain, volume, nav, cost))
    df = pd.DataFrame(result, columns=COLUMNS + ['volume', 'nav', 'cost'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.apply(verify, axis=1)
    df.drop(columns=['volume', 'cost'], inplace=True)
    return df


def huabao(datafile: str) -> pd.DataFrame:
    with open(datafile) as f:
        lines = []
        for l in f.readlines():
            # lines += re.sub('－', '-', l.rstrip('\n')).split()
            lines += l.rstrip('\n').split()
    i = 0
    while not re.match(r'\d+\.\d\d', lines[i]):
        i += 1
    asset = float(lines[i])
    i += 3
    while not re.match(r'\d+\.\d\d', lines[i]):
        i += 1
    total_mv = float(lines[i])
    cash = float(lines[i + 1])
    i += 2
    assert round(total_mv + cash, 2) == asset, \
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('华宝', 'cny', 'cash', '现金', '货币', 0, cash, 0, None, None, None)]
    while not lines[i].startswith('总仓位'):
        i += 1
    s = lines[i + 1]
    total_hg = float(re.sub('－', '-', s) if s[0] == '－' else s[1:])
    i += 1
    try:
        while lines[i] != '仓位':
            i += 1
        i += 1
        while True:
            while not re.match(r'\d{6}\.(SZ|SH)', lines[i]):
                i += 1
            name = lines[i - 4]
            cost = float(lines[i - 3])
            volume = int(lines[i - 2])
            s = lines[i - 1]
            hold_gain = float(re.sub('－', '-', s) if s[0] == '－' else s[1:])
            code = lines[i][: 6]
            if '.' in lines[i + 3]:
                market_value = float(lines[i + 2])
                nav = float(lines[i + 3])
                v2 = int(lines[i + 4])
            else:
                nav = float(lines[i + 2])
                v2 = int(lines[i + 3])
                market_value = float(lines[i + 5])
            assert v2 == volume, print("v2{} != volume{}".format(v2, volume))
            assert market_value == round(nav * volume, 2) or market_value == round(nav * volume * 10, 2), \
                print("mv({}) != nav({}) * volume({})".format(market_value, nav, volume))
            if market_value == round(nav * volume * 10, 2):
                volume *= 10
            if code[0] == '1':
                typ, risk = '转债', 2
            else:
                name, typ, risk = SECURITIES[code]
            i += 6
            result.append(('华宝', 'cny', code, name, typ, risk, market_value, hold_gain, volume, nav, cost))
    except IndexError:
        pass
    df = pd.DataFrame(result, columns=COLUMNS + ['volume', 'nav', 'cost'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.apply(verify, axis=1)
    df.drop(columns=['volume', 'cost'], inplace=True)
    return df


def huasheng(datafile: str) -> pd.DataFrame:
    s = re.sub(r'.+_', '', datafile[: -4])
    cash2 = float(s) if s else 0

    with open(datafile) as f:
        lines = []
        for line in f.readlines():
            lines += re.sub(r'[,＋]', '', re.sub('－', '-', line)).rstrip('\n').split()

    i = 0
    while not lines[i].startswith('资产净值'):
        i += 1
    currency = 'hkd' if '港币' in lines[i] else 'usd'
    while not re.match(r'^[-\d]+\.\d\d$', lines[i]):
        i += 1
    asset = float(lines[i + 1])
    i += 2
    while not re.match(r'^[.\d]+$', lines[i]):
        i += 1
    total_mv = float(lines[i])
    total_hg = float(lines[i + 2])
    i += 3
    while not re.match(r'.*[\d.]+$', lines[i]):
        i += 1
    cash = float(lines[i]) + float(lines[i + 1])
    assert round(total_mv + cash, 2) == asset, \
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('华盛', currency, 'cash', '现金', '货币', 0, cash, 0)]
    i += 15
    while not lines[i].endswith('成本价'):
        i += 1
    i += 1
    codes = []
    while len(lines) - i >= 8 and lines[i] != '行情' and lines[i + 1] != '行情':
        # name = lines[i]
        # i += 1
        while not re.match(r'^\d+$', lines[i]):
            i += 1
        volume = int(lines[i])
        i += 1
        hold_gain = float(lines[i])
        i += 1
        nav = float(lines[i])
        i += 1
        code = re.search(r'[0-9A-Z]+', lines[i]).group()
        if currency == 'hkd':
            code = code[:5]
        else:
            code = code[:4]
        i += 1
        name, type, risk = SECURITIES[code]
        while not re.match(r'^[\d.]+$', lines[i]):
            i += 1
        market_value = float(lines[i])
        i += 1
        while not re.match(r'^[\d.]+$', lines[i]):
            i += 1
        cost = float(lines[i])
        i += 1
        if code not in codes:
            codes.append(code)
            result.append(('华盛', currency, code, name, type, risk, market_value, hold_gain, volume, nav, cost))
    df = pd.DataFrame(result, columns=COLUMNS + ['volume', 'nav', 'cost'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.at[0, 'market_value'] += cash2
    df.apply(verify, axis=1)
    df.drop(columns=['volume', 'cost'], inplace=True)
    return df


def futu(datafile: str) -> pd.DataFrame:
    with open(datafile) as f:
        lines = []
        for line in f.readlines():
            line = re.sub(r'[,，]', '', re.sub(r'[+＋]', ' ', re.sub('[-－]', ' -', line)))
            lines += line.rstrip('\n').split()

    i = 0
    while not lines[i].startswith('资产净值'):
        i += 1
    currency = 'hkd' if '港币' in lines[i] else 'usd'
    asset = float(lines[i + 1])
    i += 2
    while not lines[i].startswith('持仓盈亏'):
        i += 1
    total_hg = float(lines[i + 3])
    i += 4
    while not lines[i].startswith('现金可提'):
        i += 1
    total_mv = float(lines[i + 1])
    cash = float(lines[i + 2])
    i += 3
    assert round(total_mv + cash, 2) == asset, \
        print("total_mv({}) + cash({}) != asset({})".format(total_mv, cash, asset))
    result = [('富途', currency, 'cash', '现金', '货币', 0, cash, 0)]
    i += 2
    while not lines[i].startswith('名称代码'):
        i += 1
    i += 1
    while len(lines) - i >= 7:
        while not re.match(r'^[\d]+\.\d\d$', lines[i]):
            i += 1
        market_value = float(lines[i])
        hold_gain = float(lines[i + 1])
        i += 2
        while not re.match(r'^[A-Z]{4}$', lines[i]):
            i += 1
        code = lines[i]
        i += 1
        while not re.match(r'^[\d]+$', lines[i]):
            i += 1
        volume = lines[i]
        name, type, risk = SECURITIES[code]
        nav = round(market_value / int(volume), 2)
        result.append(('富途', currency, code, name, type, risk, market_value, hold_gain, nav))
        i += 7
    df = pd.DataFrame(result, columns=COLUMNS + ['nav'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert sum_mv == asset, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert sum_hg == total_hg, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    return df


def danjuan(datafile: str) -> pd.DataFrame:
    def get_plan(plan: str) -> list:
        resp = requests.get(url + 'plan/' + plan, headers=headers)
        assert resp.status_code == 200, print('status_code({}) != 200'.format(resp.status_code))
        result = []
        for i in resp.json()['data']['items']:
            plan_code = i['plan_code']
            code = 'F' + i['fd_code']
            market_value = float(i['market_value'])
            hold_gain = float(i['hold_gain'])
            nav = float(i['nav'])
            if market_value:
                name, type, risk = SECURITIES[code]
                result.append(('蛋卷' + plan_code, 'cny', code, name, type, risk, market_value, hold_gain, nav))
        return result

    if os.path.isfile(datafile):
        return pd.read_csv(datafile, index_col=0)

    with open('auth/dj_cookie.txt', 'r') as f:
        cookie = f.read()[:-1]  # delete last '\n'
    # url = 'https://danjuanapp.com/djapi/holding/'
    url = 'https://danjuanfunds.com/djapi/holding/'
    headers = {
        'Cookie': cookie,
        'Host': url.split('/')[2],
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_1) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/88.0.4324.192 Safari/537.36'}

    response = requests.get(url + 'summary/v2', headers=headers)
    assert response.status_code == 200, print('status_code({}) != 200'.format(response.status_code))
    asset = float(response.json()['data']['total_assets'])
    total_hg = float(response.json()['data']['hold_gain'])
    result = []
    for i in response.json()['data']['items']:
        code = i['fd_code']
        if code.isnumeric():
            code = 'F' + code
        # if code in ['TIA06020', 'TIA06028']:  # 螺丝钉金钉宝主动优选, 螺丝钉金钉宝指数增强
        #     result.extend(get_plan(code))
        #     name, type, risk = i['fd_name'], '债券', 1
        # else:
        #     if code == 'TIA06019':      # 螺丝钉银钉宝365天
        #         name, type, risk = i['fd_name'], '债券', 1
        #     else:
        name, type, risk = SECURITIES[code]
        market_value = float(i['market_value'])
        hold_gain = float(i['hold_gain'])
        nav = float(i['nav'])
        result.append(('蛋卷', 'cny', code, name, type, risk, market_value, hold_gain, nav))
    df = pd.DataFrame(result, columns=COLUMNS + ['nav'])
    sum_mv = round(df['market_value'].sum(), 2)
    assert abs(sum_mv - asset) <= 1, print("sum_mv({}) != asset({})".format(sum_mv, asset))
    sum_hg = round(df['hold_gain'].sum(), 2)
    assert abs(sum_hg - total_hg) <= 1, print("sum_hg({}) != total_hg({})".format(sum_hg, total_hg))
    df.to_csv(datafile)
    return df


def tonghs(datafile: str) -> pd.DataFrame:
    if os.path.isfile(datafile):
        df = pd.read_csv(datafile, index_col=0)
        # df['code'] = df['code'].apply(lambda x: '{0:06d}'.format(x))
        return df

    with open('auth/ths_cookie.txt', 'r') as f:
        cookie = f.read()[:-1]  # delete last '\n'
    url = 'https://trade.5ifund.com/pc_query/trade_queryIncomeWjZeroList.action?_=1615356614458'
    url2 = 'https://trade.5ifund.com/pc_query/trade_currentShareList.action?_=1615344156640'
    headers = {
        'Cookie': cookie,
        'Host': url.split('/')[2],
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_2_1) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/88.0.4324.192 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest'}

    result = []
    response = requests.get(url, headers=headers)
    assert response.status_code == 200, print('status_code({}) != 200'.format(response.status_code))
    items = response.json()['singleData']['IncomeShareListResult']
    for i in items:
        code = 'F' + i['fundCode']
        name, type, risk = SECURITIES[code]
        market_value = float(i['totalVol'])
        hold_gain = float(i['sumIncome'])
        nav = 1.0
        result.append(('同花顺', 'cny', code, name, type, risk, market_value, hold_gain, nav))
    response = requests.get(url2, headers=headers)
    assert response.status_code == 200, print('status_code({}) != 200'.format(response.status_code))
    items = response.json()['singleData']['currentShareList']
    for i in items:
        code = 'F' + i['fundCode']
        name, type, risk = SECURITIES[code]
        market_value = float(i['currentValueText'])
        hold_gain = float(i['totalprofitlossText'])
        nav = float(i['navText'])
        result.append(('同花顺', 'cny', code, name, type, risk, market_value, hold_gain, nav))
    df = pd.DataFrame(result, columns=COLUMNS + ['nav'])
    df.to_csv(datafile)
    return df


def run(file: str) -> pd.DataFrame:
    platforms = {'zsb': zhaoshang_bank, 'hsb': hangseng_bank, 'yh': yinhe, 'hb': huabao, 'hs': huasheng, 'ft': futu,
                 'dj': danjuan, 'ths': tonghs}
    func = platforms[re.search(r'.*/(.*)_.*', file).group(1)]
    return func(file)


def to_cny(row: pd.Series, key: str, rates: tuple) -> float:
    h2c, u2c = rates
    value = row[key] * h2c if row['currency'] == 'hkd' else row[key] * u2c if row['currency'] == 'usd' else row[key]
    return round(value, 2)


def gain_rate(row: pd.Series) -> float:
    try:
        rate = row['hold_gain'] / (row['market_value'] - row['hold_gain'])
    except ZeroDivisionError:
        rate = 0
    # return '{:.2%}'.format(rate)
    return round(rate, 4)


def hkd_usd_rate() -> tuple:
    url = 'https://www.boc.cn/sourcedb/whpj/'
    res = requests.get(url)
    res.raise_for_status()
    res.encoding = res.apparent_encoding
    soup = BeautifulSoup(res.text, 'lxml')

    tab = soup.find_all("table")[1]
    trs = tab.find_all("tr")
    trs.pop(0)
    dic = {}
    for tr in trs:
        tds = tr.find_all("td")
        if tds[0].text == '港币':
            dic['港币'] = round(float(tds[4].text) / 100, 4)
        elif tds[0].text == '美元':
            dic['美元'] = round(float(tds[4].text) / 100, 4)
    if '港币' not in dic or '美元' not in dic:
        rates = CurrencyRates()
        dic['港币'] = round(rates.get_rate('HKD', 'CNY'), 4)
        dic['美元'] = round(rates.get_rate('USD', 'CNY'), 4)
    return dic['港币'], dic['美元']


def fill(rates: tuple, df: pd.DataFrame) -> pd.DataFrame:
    h2c, u2c = rates
    df['name'] = df['name'].apply(lambda s: s if len(s) <= 10 else s[: 8] + '..')  # truncate name
    df['market_value'] = df.apply(lambda row: to_cny(row, 'market_value', (h2c, u2c)), axis=1)
    df['hold_gain'] = df.apply(lambda row: to_cny(row, 'hold_gain', (h2c, u2c)), axis=1)
    df['gain_rate'] = df.apply(gain_rate, axis=1)
    col2 = ['platform', 'currency', 'code', 'name', 'type', 'risk', 'nav', 'market_value', 'hold_gain', 'gain_rate']
    return df.reindex(columns=col2)


def run_all(rates: tuple, files: list) -> pd.DataFrame:
    platforms = ['zsb_', 'hsb_', 'yh_', 'hb_', 'hs_', 'ft_', 'dj_', 'ths_']   # to sort files
    fs = sorted(files)
    frames = []
    for p in platforms:
        for f in fs:
            if os.path.basename(f).startswith(p):
                # print(f, p)
                frames.append(run(f))
    df = pd.concat(frames)
    df = fill(rates, df)
    return df


def to_execl(xlsx: str, rates: tuple, sheet: str, df: pd.DataFrame):
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()

    ws = wb[sheet]
    last_row = ws.max_row
    last_col = ws.max_column

    for i in range(2, last_row + 1):
        for j in range(6, last_col):
            ws.cell(row=i, column=j).number_format = '#,##,0.00'
        ws.cell(row=i, column=last_col).number_format = '0.00%'

    h2c, u2c = rates
    ws.cell(row=last_row + 1, column=1).value = h2c
    ws.cell(row=last_row + 1, column=2).value = u2c

    summaries = [
        {'location': (last_row + 2, 1),
         'letter': 'A',
         # 'labels': ['招商银行', '恒生银行', '银河', '华宝', '华盛*', '富途*', '蛋卷*', '同花顺'],
         'labels': ['招商银行', '恒生银行', '银河', '华盛*', '富途*', '蛋卷*'],
         'category': 'platform',
         'anchor': 'K1'},
        {'location': (last_row + 2, 4),
         'letter': 'B',
         'labels': ['cny', 'hkd', 'usd'],
         'category': 'currency',
         'anchor': 'K16'},
        {'location': (last_row + 6, 4),
         'letter': 'F',
         'labels': [0, 1, 2, 3],
         'category': 'risk',
         'anchor': 'K31'},
        {'location': (last_row + 2, 7),
         'letter': 'E',
         'labels': ['货币', '债券', '转债', '指数', '主动', '股票', '医药', '中概股'],
         'category': 'type',
         'anchor': 'K46'}
    ]

    for summary in summaries:
        row, col = summary['location']
        le = [get_column_letter(j) for j in range(col, col + 3)]
        for i in range(len(summary['labels'])):
            ws.cell(row=row + i, column=col).value = summary['labels'][i]
            c = ws.cell(row=row + i, column=col + 1)
            c.number_format = "#,##,0.00"
            c.value = '=SUMIF(${0}$2:${0}${1},{2}{3},$H$2:$H${1})'.format(summary['letter'],
                                                                          last_row, le[0], row + i)
            c = ws.cell(row=row + i, column=col + 2)
            c.number_format = "#,##,0.00"
            c.value = '=SUMIF(${0}$2:${0}${1},{2}{3},$I$2:$I${1})'.format(summary['letter'],
                                                                          last_row, le[0], row + i)
        ws.cell(row=row + i + 1, column=col).value = 'sum'
        c = ws.cell(row=row + i + 1, column=col + 1)
        c.number_format = "#,##,0.00"
        c.value = '=SUM({0}{1}:{0}{2})'.format(le[1], row, row + i)
        c = ws.cell(row=row + i + 1, column=col + 2)
        c.number_format = "#,##,0.00"
        c.value = '=SUM({0}{1}:{0}{2})'.format(le[2], row, row + i)

        if summary['letter'] == 'E':
            for j in range(len(summary['labels']) + 1):
                c = ws.cell(row=row + j, column=col + 3)
                c.number_format = '0.00%'
                c.value = '={1}{0}/({2}{0}-{1}{0})'.format(row + j, le[2], le[1])

        pie = PieChart()
        labels = Reference(ws, min_col=col, min_row=row, max_row=row + i)
        data = Reference(ws, min_col=col + 1, min_row=row - 1, max_row=row + i)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = summary['category']
        ws.add_chart(pie, summary['anchor'])
    wb.save(xlsx)


def to_mysql(date_str: str, df: pd.DataFrame):
    total = df['market_value'].sum()
    df = df.drop(columns=['gain_rate'])
    dic = {k: 'first' for k in df.columns}
    dic['market_value'] = dic['hold_gain'] = 'sum'
    df = df.groupby('code', as_index=False).agg(dic)
    df = df.sort_values(by=['platform'])

    df['date'] = datetime.strptime(date_str, '%y%m%d')
    assert round(total, 2) == round(df['market_value'].sum(), 2)
    print(df)
    db = MySql(database='portfolio')
    db.from_frame('asset', df)


def excel_mysql(xlsx: str):
    db = MySql(database='portfolio')
    book = load_workbook(xlsx, data_only=True)
    for sheet in book.worksheets:
        columns = []
        items = []
        total_market_value = 0
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
            cells = [cell.value for cell in row]
            if cells[0] == 'platform':
                columns = cells[: -1].copy()
                continue
            if cells[0] == 'sum':
                total_market_value = cells[1]
                continue
            if cells[1] not in ['rmb', 'cny', 'hkd', 'usd']:
                continue
            if (cells[0] in ['恒生银行', '同花顺'] or cells[0].startswith('蛋卷')) and cells[2].isnumeric():
                cells[2] = 'F' + cells[2]
            if cells[0] not in ['华宝', '银河'] or cells[2][0] not in ['1', '7']:
                if cells[2] == 'product':
                    cells[2] = cells[3][4:] if cells[3].startswith('招银理财') else cells[3]
                    if cells[2].startswith('招智睿远..'):
                        cells[2] = '招智睿远平衡二十七期'
                    elif cells[2].startswith('招赢日日盈'):
                        cells[2] = '招赢日日盈'
                cells[3], cells[4], cells[5] = SECURITIES[cells[2]]
            items.append(cells[: -1])

        df = pd.DataFrame(items, columns=columns)
        # duplicates = df[df.duplicated(['code'], keep=False)]

        names = ['蛋卷CSI666', '蛋卷CSI1033', '蛋卷TIA06020', '蛋卷TIA06028']
        items = []
        for name in names:
            df1 = df[df['platform'] == name]
            if not df1.empty:
                code = name[2:]
                market_value = round(df1['market_value'].sum(), 2)
                hold_gain = round(df1['hold_gain'].sum(), 2)
                # gain_rate = round(hold_gain / market_value, 4)
                items.append(['蛋卷', 'cny', code, *SECURITIES[code], np.nan, market_value, hold_gain])
        df2 = pd.DataFrame(items, columns=columns)
        df = df.loc[~df['platform'].isin(names)]
        df = pd.concat([df, df2], ignore_index=True)

        dic = {k: 'first' for k in columns}
        dic['market_value'] = dic['hold_gain'] = 'sum'
        df = df.groupby('code', as_index=False).agg(dic)
        df = df.sort_values(by=['platform'])

        df['date'] = datetime.strptime(sheet.title, '%y%m%d')
        assert round(total_market_value, 2) == round(df['market_value'].sum(), 2)
        print(df)
        db.from_frame('asset', df)
    return


def main():
    if len(sys.argv) < 2:
        print('Usage: {} txt'.format(sys.argv[0]))
        print('       {} date_str [xlsx]'.format(sys.argv[0]))
        print('       {} xlsx'.format(sys.argv[0]))
        sys.exit(1)

    path = sys.argv[1]
    if path.endswith('.txt'):
        if not os.path.isfile(path):
            with open(path, 'w') as fp:
                fp.write(pyperclip.paste())
        print(run(path))
    elif path.endswith('.csv'):
        print(run(path))
    elif os.path.isdir(path):
        rates = hkd_usd_rate()
        path = path.rstrip('/')
        df = run_all(rates, [os.path.join(path, file) for file in os.listdir(path)])
        if len(sys.argv) == 2:
            to_mysql(path, df)
        else:
            to_execl(sys.argv[2], rates, path, df)
    elif path.endswith('.xlsx'):
        excel_mysql(path)
    else:
        assert False


if __name__ == "__main__":
    main()
