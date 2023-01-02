#! /usr/bin/python3
import os
import re
import sys
from datetime import datetime
from pprint import pprint
import time

import pandas as pd
import pyperclip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from thresholds import COLUMNS, THRESHOLDS, name_code
from mysql import MySql
from mongo import Mongo
from xueqiu import Xueqiu

INDEXES = [
    # PB
    '银行行业', '地产行业', '证券行业', '军工行业', '环保行业', '基建行业', '建筑材料',
    # EP
    '上证红利', '50AH优选', '基本面50', '央视50', '中证红利', '300价值', '上证50', '上证180', 'H股指数', '恒生指数',
    # PE
    '中证100', '沪深300', '中证500', '500增强', '500低波动', '中证800', '中证1000', '红利机会', '中证养老',
    '创业板', '深证100', '医药100', '中证医疗', '生物科技', '科技100', '中证消费', '中证白酒', '食品饮料',
    '可选消费', '消费红利', '消费龙头', '消费50', '沪港深消费50', '深证成指', '基本面60', '基本面120', '深红利',
    '纳斯达克100', '标普500', '标普科技', '美股消费', '全球医疗', '香港中小', '中概互联', '恒生科技',
    'A股龙头', '竞争力指数', 'MSCI质量', '科创50', '家用电器',
    # other
    '美国房地产', '十年期国债', '10年期国债（A股）', '10年期国债（美股）'
]


def parse_fund_code(file: str) -> list:
    with open(file) as f:
        lines = []
        for l in f.readlines():
            lines += l.rstrip('\n').split()

    reg_date = re.compile(r'20\d{6}')
    result = []
    try:
        i = 0
        # while True:
        while reg_date.search(lines[i]) is None:
            i += 1
        date = reg_date.search(lines[i]).group()
        # print(date)
        i += 1
        while True:
            while lines[i] not in INDEXES and not lines[i].startswith('中概互联') \
                    and not lines[i].startswith('恒生科技'):
                if lines[i].startswith('永续A') or lines[i].startswith('注'):
                    i += 1
                    break
                i += 1
            else:
                if lines[i].startswith('中概互联'):
                    key = '中概互联'
                elif lines[i].startswith('恒生科技'):
                    key = '恒生科技'
                else:
                    key = lines[i]
                if key == '中概互联':
                    if lines[i + 1] == '513050':            # there is no value
                        i += 2
                        continue
                    if lines[i + 1].startswith('市销率'):
                        i += 1
                elif key == '十年期国债':
                    key = '10年期国债（A股）'
                i += 1
                while re.search(r'\d{6}', lines[i]) is None:
                    i += 1
                on, off = lines[i], None
                i += 1
                if re.search(r'\d{6}', lines[i]):
                    off = (lines[i])
                    i += 1
                elif key not in ['上证红利', '科创50']:
                    off, on = on, None
                # print(key, on, off)
                result.append((key, on, off))
    except IndexError:
        pass
    return result


def save_thresholds(file: str):
    df = pd.DataFrame(THRESHOLDS, columns=COLUMNS)
    df2 = pd.DataFrame(parse_fund_code(file), columns=['_id', '场内代码', '场外代码'])
    df = pd.merge(df, df2, on='_id')
    print(df)
    # Mongo().save('threshold', df)


def parse_valuations(file: str) -> list:
    if not os.path.isfile(file):
        with open(file, 'w') as fp:
            fp.write(pyperclip.paste())

    with open(file) as f:
        lines = []
        for l in f.readlines():
            lines += l.rstrip('\n').split()

    reg_date = re.compile(r'20\d{6}')
    result = []
    try:
        i = 0
        while True:
            while reg_date.search(lines[i]) is None:
                i += 1
            date = reg_date.search(lines[i]).group()
            # print(date)
            dic = {'_id': date}
            i += 1
            while True:
                while lines[i] not in INDEXES and not lines[i].startswith('中概互联') \
                        and not lines[i].startswith('恒生科技'):
                    if lines[i].startswith('永续A') or lines[i].startswith('注'):
                        i += 1
                        break
                    i += 1
                else:
                    if lines[i].startswith('中概互联'):
                        key = '中概互联'
                    elif lines[i].startswith('恒生科技'):
                        key = '恒生科技'
                    else:
                        key = lines[i]
                    if key == '中概互联':
                        if lines[i + 1] == '513050':            # there is no value
                            i += 2
                            continue
                        if lines[i + 1].startswith('市销率'):
                            i += 1
                    elif key == '十年期国债':
                        key = '10年期国债（A股）'
                    i += 1
                    while lines[i] == '*':
                        i += 1
                    value = re.search(r'[\d.]+', lines[i]).group(0)
                    if key == '300价值' and value == lines[i]:
                        value = round(100 / float(value), 2)        # as 300价值 used PE before 20100225
                    else:
                        value = float(value)
                    i += 1
                    # print(key, value)
                    dic[key] = value
                    continue
                break               # break out of multiple loops as encountered '注：'
            result.append(dic)
    except IndexError:
        pass
    return result


def check(valuations: list):
    cur = {'_id'}
    for valuation in valuations:
        date = valuation['_id']
        s = set(valuation.keys())
        if cur == s:
            continue
        a = s - cur
        if a:
            print("{} add {}".format(date, a))
            cur |= a
        d = cur - s
        if d:
            print("{} missing {}".format(date, d))


def save_valuations(file: str):
    valuations = parse_valuations(file)
    check(valuations)
    # pprint(valuations)
    # for date, valuation in valuations:
    #     if '300价值' in valuation:
    #         print(date, valuation['300价值'])

    print(len(valuations))
    df = pd.DataFrame(valuations)
    df['_id'] = pd.to_datetime(df['_id'])
    print(df)
    # Mongo().save('valuation', df)


def update_valuations(xlsx: str):
    mongo = Mongo()
    ms = mongo.find_last('valuation')['_id']
    start = datetime.fromtimestamp(ms / 1000.0).strftime('%Y%m%d')

    wb = load_workbook(xlsx)
    wss = [ws.title for ws in wb.worksheets]
    excel = pd.ExcelFile(xlsx)
    valuations = []
    for ws in wss:
        if not re.search(r'20\d{6}$', ws) or ws <= start:
            continue
        df = pd.read_excel(excel, ws)
        dic = dict(zip(df['名称'].tolist(), df[ws].tolist()))
        dic['_id'] = ws
        valuations.append(dic)

    print(len(valuations))
    if not valuations:
        return
    # the valuation hasn't ('中证500', '美国房地产', '10年期国债（美股）', '10年期国债（A股）', '消费50')
    df = pd.DataFrame(valuations)
    df['_id'] = pd.to_datetime(df['_id'])
    print(df)
    mongo.save('valuation', df)


def to_lowest(row: pd.Series) -> float:
    val, lowest, low = row.iat[7], row['最低'], row['低估']
    if row['参考指标'] == '盈利收益率':
        if val < low:
            return None
    elif val > low:
        return None
    return round((lowest - val) / (lowest - low), 4)


def to_lowest2(row: pd.Series) -> float:
    val, lowest, low = row.iat[7], row['最低'], row['低估']
    if row['参考指标'] == '盈利收益率':
        if val < low:
            return None
        return round((val - lowest) / val, 4)
    if val > low:
        return None
    return round((lowest - val) / val, 4)


def to_low(row: pd.Series) -> float:
    val, low, high = row.iat[7], row['低估'], row['高估']
    if row['参考指标'] == '盈利收益率':
        if val >= low or val <= high:
            return None
    elif val <= low or val >= high:
        return None
    return round((low - val) / (low - high), 4)


def to_low2(row: pd.Series) -> float:
    val, low, high = row.iat[7], row['低估'], row['高估']
    if row['参考指标'] == '盈利收益率':
        if val <= high:
            return None
        return round((val - low) / val, 4)
    if val >= high:
        return None
    return round((low - val) / val, 4)


def to_high(row: pd.Series) -> float:
    val, high, highest = row.iat[7], row['高估'], row['最高']
    if row['参考指标'] == '盈利收益率':
        if val > high:
            return None
    elif val < high:
        return None
    return round((high - val) / (high - highest), 4)


def to_high2(row: pd.Series) -> float:
    val, low, high = row.iat[7], row['低估'], row['高估']
    if row['参考指标'] == '盈利收益率':
        if val >= low:
            return None
        return round((val - high) / val, 4)
    if val <= low:
        return None
    return round((high - val) / val, 4)


def to_highest2(row: pd.Series) -> float:
    val, high, highest = row.iat[7], row['高估'], row['最高']
    if row['参考指标'] == '盈利收益率':
        if val >= high:
            return None
        return round((val - highest) / val, 4)
    if val <= high:
        return None
    return round((highest - val) / val, 4)


def get_valuation_with_threshold(dic: dict) -> pd.DataFrame:
    columns = ['代码', '名称', '参考指标', '最低', '低估', '高估', '最高']
    df = pd.DataFrame(THRESHOLDS, columns=columns)
    date = dic['_id']

    df2 = pd.DataFrame(dic.items(), columns=['名称', date])
    # print(len(df), len(df2))
    # df = pd.merge(df, df2, on='名称', how='left', indicator=True)
    # print(df[df['_merge'] == 'left_only'])
    df = pd.merge(df, df2, on='名称', how='left')
    nan = df[df[date].isna()]
    assert len(nan) == 0, print(nan)

    df['距最低'] = df.apply(to_lowest, axis=1)     # to_lowest(), to_low(), to_high() are just for sort
    df['距低估'] = df.apply(to_low, axis=1)
    df['距高估'] = df.apply(to_high, axis=1)
    df = df.sort_values(by=['距最低', '距低估', '距高估'])
    df['距最低'] = df.apply(to_lowest2, axis=1)
    df['距低估'] = df.apply(to_low2, axis=1)
    df['距高估'] = df.apply(to_high2, axis=1)
    df['距最高'] = df.apply(to_highest2, axis=1)
    return df.reset_index(drop=True)


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    wb = load_workbook(xlsx)
    ws = wb.copy_worksheet(wb.worksheets[-1])       # copy a old sheet as template to avoid adjust size
    ws.title = sheet
    wb.active = len(wb.worksheets) - 1

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()

    wb = load_workbook(xlsx)
    ws = wb[sheet]
    last_row = ws.max_row
    last_col = ws.max_column
    for i in range(2, last_row + 1):
        for j in range(4, 9):
            ws.cell(row=i, column=j).number_format = '#,##,0.00'
        for j in range(9, last_col + 1):
            ws.cell(row=i, column=j).number_format = '0.00%'
    colors = ['99CC00', 'FFCC00', 'FF6600']
    i = 1
    for k in range(9, 12):
        fill = PatternFill(patternType='solid', fgColor=colors[k - 9])
        while ws.cell(row=i, column=k).value is not None:
            for j in range(1, last_col + 1):
                ws.cell(row=i, column=j).fill = fill
            i += 1

    wb.save(xlsx)


def get_valuation_with_star(dic: dict, star: float) -> dict:
    date = dic['_id']
    dic = {v: dic[k] for k, v in name_code().items()}

    snowball = Xueqiu()
    dic2 = snowball.last_close('sh000985')
    assert date == dic2['date'].replace('-', '')
    dic['timestamp'] = int(time.mktime(time.strptime(date, "%Y%m%d")))
    dic['sh000985'] = dic2['中证全指']
    dic['star'] = star
    return dic


def to_mysql(dic: dict):
    db = MySql(database='portfolio')
    last = db.last_row('valuation', 'timestamp')
    # print(last)
    # for key in dic:
    #     if dic[key] != last[key]:
    #         print(f"Different value for {key}: {dic[key]} != {last[key]}")
    if dic['timestamp'] > last['timestamp']:
        db.insert('valuation', dic)
        print('inserted row(%s) after row(%s)' %
              (datetime.fromtimestamp(dic['timestamp']).strftime('%y%m%d'),
               datetime.fromtimestamp(last['timestamp']).strftime('%y%m%d')))


def usage():
    print('Usage: %s star txt' % sys.argv[0])       # update spreadsheet from clipboard or file
    print('       %s star txt xlsx' % sys.argv[0])  # dump to spreadsheet and database
    print('       %s xlsx' % sys.argv[0])           # update db from spreadsheet
    sys.exit(1)


def main():
    # save_thresholds(sys.argv[1])
    # save_valuations(sys.argv[1])

    if len(sys.argv) > 2:
        if not sys.argv[2].endswith('.txt'):
            usage()
        dic = parse_valuations(sys.argv[2])[0]
        date = dic['_id']
        val_df = get_valuation_with_threshold(dic)
        print(val_df)
        val_dict = get_valuation_with_star(dic, float(sys.argv[1]))
        pprint(val_dict)
        if len(sys.argv) > 3:
            if not sys.argv[3].endswith('.xlex'):
                usage()
            to_excel(sys.argv[2], date, val_df)
            to_mysql(val_dict)
        print('%s: 中证全指 %.2f 星级 %.1f' % (date, val_dict['sh000985'], val_dict['star']))
    elif sys.argv[1].endswith('.xlsx'):
        update_valuations(sys.argv[1])
    else:
        usage()


if __name__ == "__main__":
    main()
