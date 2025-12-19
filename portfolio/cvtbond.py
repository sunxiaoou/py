#! /usr/local/bin/python3
import re
import sys
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from excel_tool import duplicate_last_sheet, df_to_sheet
from mysql import MySql
from snowball import Snowball

pd.set_option('display.max_rows', 200)
pd.set_option('display.max_columns', 100)

MAX_COLUMN = 29


def get_bonds(xlsx: str) -> (dict, pd.DataFrame):
    wb = load_workbook(xlsx, data_only=True)

    ws = wb['强赎预警']
    dic = {}
    for j in range(1, ws.max_column):
        title = ws.cell(row=1, column=j).value
        if title in ['代码', '强赎天计数']:
            dic[title] = j
            if(len(dic)) == 2:
                break
    warning = {}
    for i in range(2, ws.max_row + 1):
        code = ws.cell(row=i, column=dic['代码']).value
        if not isinstance(code, int):
            break
        warning[str(code)] = ws.cell(row=i, column=dic['强赎天计数']).value

    # ws = wb['今天禄得']
    ws = wb['今天排名']
    data = []
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row - 2, max_col=MAX_COLUMN):
        cells = [cell.value for cell in row]
        if not cells[0]:
            break
        data.append(cells)
    titles = [i for i in data[0] if i]
    title = titles[0]
    for i in range(1, len(data[0])):
        if data[0][i] is None:
            data[0][i] = title
        else:
            title = data[0][i]
    frame = pd.DataFrame(data[2:], columns=[data[0], data[1]])
    base_cols = ['代码', '名称', '价格', '涨幅', '强赎提示']
    base1 = frame['无阈值'][base_cols]
    base2 = frame['150元阈值'][base_cols]
    base3 = frame['浮动阈值'][base_cols]
    base4 = frame['130元阈值'][base_cols]
    df = pd.concat([base1, base2, base3, base4], ignore_index=True) \
        .drop_duplicates(subset='代码', keep='first')
    rank1 = frame['无阈值'][['代码', '最新排名']].rename(columns={'最新排名': '无阈值排名'})
    rank2 = frame['150元阈值'][['代码', '最新排名']].rename(columns={'最新排名': '150阈值排名'})
    rank3 = frame['浮动阈值'][['代码', '最新排名']].rename(columns={'最新排名': '浮动阈值排名'})
    rank4 = frame['130元阈值'][['代码', '最新排名']].rename(columns={'最新排名': '130阈值排名'})
    df = df.merge(rank1, on='代码', how='left')
    df = df.merge(rank2, on='代码', how='left')
    df = df.merge(rank3, on='代码', how='left')
    df = df.merge(rank4, on='代码', how='left')
    df = df.rename(columns={'涨幅': '涨跌幅%', '强赎提示': '强赎天计数'})
    df['代码'] = df['代码'].apply(lambda x: str(x))
    df['涨跌幅%'] = df['涨跌幅%'].apply(lambda x: float(x) * 100)
    # print(df)
    return warning, df


def get_my_list(xlsx: str) -> pd.DataFrame:
    wb = load_workbook(xlsx, data_only=True)
    ws = wb.worksheets[-1]
    lst = []
    for i in range(1, ws.max_row):
        try:
            code = ws.cell(row=i, column=3).value
            name = ws.cell(row=i, column=4).value
            if code[0] in ['1', '7'] and name[2] == '转':
                lst.append((code, name))
        except TypeError:
            pass
    df = pd.DataFrame(lst, columns=['代码', '名称']).drop_duplicates()
    return df


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    duplicate_last_sheet(xlsx, sheet, df.shape[0])
    df_to_sheet(df, xlsx, sheet, overlay=True, header=True)


def update_price(codes: list, db: MySql):
    table = 'cvtbond_daily'
    snowball = Snowball()
    for code in codes:
        dic = db.last_row(table, 'date', 'code = "%s"' % code)
        if not dic:
            df = snowball.get_data(code)
            dic['name'] = snowball.get_name(code)
        else:
            begin_date = (dic['date'] + pd.Timedelta(days=1)).strftime('%Y-%m-%d')
            df = snowball.get_data(code, begin_date)
        if df.empty:
            break           # maybe continue
        df['code'] = code
        df['name'] = dic['name']
        df = df[['date', 'code', 'name', 'open', 'high', 'low', 'close', 'volume']]
        print(df)
        db.from_frame(table, df)
        # time.sleep(0.2)


def update_rank(df: pd.DataFrame, db: MySql):
    table = 'cvtb_rank_daily'
    dic = db.last_row(table, 'date')
    if dic['date'] < df['date'].iloc[0]:
        db.from_frame(table, df)


def main():
    if len(sys.argv) < 2:
        print('Usage: {} ref_rank_list.xlsx [rank130] [out_xlsx]'.format(sys.argv[0]))
        sys.exit(1)

    in_xlsx = sys.argv[1]
    warning, bonds = get_bonds(in_xlsx)
    out_xlsx = None
    title = '无阈值排名'
    if len(sys.argv) > 3:
        assert sys.argv[3].endswith('.xlsx')
        out_xlsx = sys.argv[3]
    elif len(sys.argv) > 2:
        if sys.argv[2].endswith('.xlsx'):
            out_xlsx = sys.argv[2]
        else:
            rank130 = int(sys.argv[2])
            df = bonds.loc[bonds['130排名'] == rank130]
            r = df.iloc[0, df.columns.get_loc('无阈值排名')]
            bonds = bonds[bonds['无阈值排名'] <= r]
            title += '(截止到<130的第{}名)'.format(rank130)
    out = bonds[['代码', '名称', '价格', '涨跌幅%', '无阈值排名', '150阈值排名', '浮动阈值排名', '130阈值排名', '强赎天计数']]
    date = re.search(r'\d{8}', in_xlsx)[0]
    print(date)
    if out_xlsx:
        to_excel(out_xlsx, date, out)

    mask = bonds['价格'] <= 170
    bonds['170排名'] = np.nan
    bonds.loc[mask, '170排名'] = range(1, mask.sum() + 1)
    bonds = bonds[['代码', '名称', '价格', '涨跌幅%', '170排名', '强赎天计数']]

    mine = Snowball().my_cvt_bonds()
    owned = pd.merge(bonds, mine[['代码']], on=['代码'])
    print('已持有的上榜转债({})'.format(len(owned)))
    print(owned)

    unowned = pd.concat([bonds, owned]).drop_duplicates(keep=False)
    unowned = unowned[unowned['价格'] < 170]
    print('未持有的<170的上榜转债({})'.format(len(unowned)))
    print(unowned)

    to_sell = pd.concat([owned[['代码']], mine[['代码']]]).drop_duplicates(keep=False)
    print('持有的未上榜转债({})'.format(len(to_sell)))
    if to_sell.shape[0]:
        to_sell = pd.merge(to_sell, mine, on=['代码'], how='left')
        to_sell['强赎天计数'] = to_sell.apply(lambda x: warning[x['代码']] if x['代码'] in warning else np.nan, axis=1)
        print(to_sell)

    owned['状态'] = 'owned'
    unowned['状态'] = 'unowned'
    if to_sell.shape[0]:
        to_sell['状态'] = 'to_sell'
        df = pd.concat([owned, unowned, to_sell[['代码', '名称', '价格', '涨跌幅%', '强赎天计数', '状态']]])
    else:
        df = pd.concat([owned, unowned])
    df['代码'] = df['代码'].apply(lambda x: 'SH' + x if x.startswith('11') else 'SZ' + x)
    codes = df['代码'].tolist()
    df['date'] = datetime.strptime(date, '%Y%m%d')
    df = df[['date', '代码', '名称', '170排名', '强赎天计数', '状态']]
    df = df.rename({'代码': 'code', '名称': 'name', '170排名': 'rank_170', '强赎天计数': 'redeem_days', '状态': 'status'},
                   axis=1)
    db = MySql(database='portfolio')
    update_price(codes, db)
    update_rank(df, db)


if __name__ == "__main__":
    main()
