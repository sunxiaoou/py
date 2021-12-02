#! /usr/bin/python3
import sys
from datetime import date
from datetime import datetime

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from openpyxl import load_workbook

from mongo import Mongo

# pd.set_option('display.max_rows', 1000)
# pd.set_option('display.max_columns', 6)


def get_close_price(code: str, begin: date) -> tuple:
    mongo = Mongo()
    name = mongo.load_info(code)['name']
    name = '{}({})'.format(name if len(name) <= 10 else name[: 8] + '..', code)
    base = mongo.load_close_price('sh510310')[['date']]         # use '沪深300ETF' as base
    base['date'] = pd.to_datetime(base['date'])
    base = base[base['date'] > begin]
    df = mongo.load_close_price(code)
    df = pd.merge(base, df, on='date', how='outer')
    df.fillna(method='ffill', inplace=True)                     # fill NaN with previous value
    df.fillna({'close': 1.0}, inplace=True)                     # no previous value, fill with 1.0
    return name, df


def xirr(df: pd.DataFrame, date_column: str, amount_column: str) -> float:
    residual = 1
    step = 0.05
    guess = 0.05
    epsilon = 0.0001
    limit = 10000

    df = df.sort_values(by=date_column).reset_index(drop=True)
    df['years'] = df.apply(lambda x: (x[date_column] - df[date_column][0]).days / 365, axis=1)
    while abs(residual) > epsilon and limit > 0:
        limit -= 1
        residual = df.apply(lambda x: x[amount_column] / pow(guess, x['years']), axis=1).sum()
        if abs(residual) > epsilon:
            if residual > 0:
                guess += step
            else:
                guess -= step
                step /= 2.0
    return guess - 1


def to_excel(xlsx: str, sheet: str, df: pd.DataFrame):
    try:
        wb = load_workbook(xlsx)
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    writer = pd.ExcelWriter(xlsx, engine='openpyxl')
    writer.book = wb
    writer.sheets = {worksheet.title: worksheet for worksheet in wb.worksheets}
    df.to_excel(writer, sheet_name=sheet, index=False)
    writer.save()


def loop_back(code: str, begin: date) -> tuple:
    name, df = get_close_price(code, begin)
    weekday = 1                                 # choose Tuesday, Mon: 0, Tue: 1, ... Sun: 6
    df = df[(df['date'] > begin) & (df['date'].dt.dayofweek == weekday)]

    df['每期定投金额'] = 1000
    df['累计定投金额'] = df['每期定投金额'].cumsum()
    fee_rate = 0.001
    df['每期持仓数量'] = df['每期定投金额'] / df['close'] * (1 - fee_rate)
    df['累计持仓数量'] = df['每期持仓数量'].cumsum()
    df['累计持仓净值'] = df['close'] * df['累计持仓数量']
    df['现金流'] = - df['每期定投金额']
    m = df.shape[0]     # number of row
    cumulative_amount = df.iloc[m - 1, df.columns.get_loc('累计定投金额')]
    cumulative_net = df.iloc[m - 1, df.columns.get_loc('累计持仓净值')]
    hold_gain = cumulative_net - cumulative_amount
    df.iloc[m - 1, df.columns.get_loc('现金流')] += cumulative_net
    return_rate = xirr(df, 'date', '现金流')
    df = df.rename({'累计持仓净值': name}, axis=1).set_index('date')
    df.index.name = None
    # print(df)
    return (name, cumulative_amount, cumulative_net, hold_gain, return_rate), df[['累计定投金额', name]]


def comparision(typ: str, codes: list, begin: date):
    results = [loop_back(i, begin) for i in codes]
    columns = ['name', '累计定投金额(万)', '累计持仓净值(万)', '累计收益(万)', '内部收益率(%)']
    df = pd.DataFrame([r[0] for r in results], columns=columns)
    df[columns[1]] = df[columns[1]].apply(lambda x: round(x / 10000))
    df[columns[2]] = df[columns[2]].apply(lambda x: round(x / 10000))
    df[columns[3]] = df[columns[3]].apply(lambda x: round(x / 10000))
    df[columns[4]] = df[columns[4]].apply(lambda x: round(x * 100))
    df.set_index('name', inplace=True)
    df.index.name = None
    print(df)
    df2 = pd.concat([r[1] for r in results], axis=1)
    df2 = df2.loc[:, ~df2.columns.duplicated()]        # remove duplicated '累计定投金额'
    print(df2)

    _, axes = plt.subplots(nrows=2, ncols=1)
    title = typ + '定投回测：累计持仓曲线及收益率'
    df2.plot(ax=axes[0], figsize=(12, 8), grid=True, rot=0, title=title)
    ax = df.plot.bar(ax=axes[1], figsize=(12, 8), rot=20)
    for i in ax.containers:
        ax.bar_label(i)
    plt.figtext(.9, .1, '- 同光和尘')
    plt.show()


def main():
    if len(sys.argv) < 3:
        print('Usage: {} code YYYYmmdd'.format(sys.argv[0]))
        sys.exit(1)

    begin = datetime.strptime(sys.argv[2], '%Y%m%d')    # .date()
    # comparision('宽基指数', ["sh502000", "sh510310", "sh510710", "f050025", "f040046", "f000071"], begin)
    # comparision('宽基指数', ["sz161039", "sh501050", "f006341", "f003318", "f090010", "f519671"], begin)
    # comparision('行业指数', ["f000248", "f162412", "f501009", "sz164906", "sh512000", "sh512800"], begin)
    # comparision('主动基金', ["f001643", "f001717", "f001810", "f005267", "f161005", "f163406"], begin)
    comparision('主动基金', ["f000595", "f001766", "f001974", "f005267", "f377240", "f540003"], begin)


if __name__ == "__main__":
    main()
