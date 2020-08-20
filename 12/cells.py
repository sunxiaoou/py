#! /usr/bin/python3


import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
# sheet = wb.get_sheet_by_name('Sheet1')
sheet = wb['Sheet1']

print(sheet['A1'])
# <Cell Sheet1.A1>

print(sheet['A1'].value)
# datetime.datetime(2015, 4, 5, 13, 34, 2)

c = sheet['B1']
print(c.value)
# 'Apples'

print('Row {}, Column {} is {}'.format(str(c.row), c.column, c.value))
# 'Row 1, Column B is Apples'

print('Cell ' + c.coordinate + ' is ' + c.value)
# 'Cell B1 is Apples'

print(sheet['C1'].value)
# 73

print(sheet.cell(row=1, column=2))
# <Cell Sheet1.B1>

print(sheet.cell(row=1, column=2).value)
# 'Apples'

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
# 1 Apples
# 3 Pears
# 5 Apples
# 7 Strawberries

print(sheet.max_row)
# 7
print(sheet.max_column)
# 3
