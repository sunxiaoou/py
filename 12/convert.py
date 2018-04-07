#! /usr/bin/python3


import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string

print(get_column_letter(1))
# 'A'

print(get_column_letter(2))
# 'B'

print(get_column_letter(27))
# 'AA'

print(get_column_letter(900))
# 'AHP'

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

print(get_column_letter(sheet.max_column))
# 'C'

print(column_index_from_string('A'))
# 1

print(column_index_from_string('AA'))
# 27
